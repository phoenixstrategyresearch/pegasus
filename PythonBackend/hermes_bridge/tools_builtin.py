"""
Built-in tools for the Pegasus agent.
These register themselves with the global registry at import time.

Tools included:
- file_read / file_write / file_list: filesystem ops (sandboxed)
- shell_exec: run shell commands (sandboxed to app container)
- web_fetch: fetch a URL and return content
- python_exec: execute Python code snippets
- pip_install: install Python packages from PyPI (no pip/subprocess needed)
- memory_read / memory_write: persistent agent memory
- skills_list / skill_view / skill_create: skill management
"""

import os
import sys
import json
import ssl
import urllib.request
import traceback
import zipfile
import io
import tempfile

# Set by AgentRunner to indicate cloud vs local mode
_CLOUD_MODE = False

# Embedded Python on iOS has no CA certificate bundle - create a permissive context
_ssl_ctx = ssl.create_default_context()
_ssl_ctx.check_hostname = False
_ssl_ctx.verify_mode = ssl.CERT_NONE
import re as _re
from pathlib import Path

# Directory for user-installed packages
_PACKAGES_DIR = os.path.join(
    os.environ.get("PEGASUS_DATA_DIR", os.path.expanduser("~/Documents/pegasus_data")),
    "packages"
)
os.makedirs(_PACKAGES_DIR, exist_ok=True)
if _PACKAGES_DIR not in sys.path:
    sys.path.insert(0, _PACKAGES_DIR)

# Directory for agent-created custom Python packages
_CUSTOM_PACKAGES_DIR = os.path.join(
    os.environ.get("PEGASUS_DATA_DIR", os.path.expanduser("~/Documents/pegasus_data")),
    "custom_packages"
)
os.makedirs(_CUSTOM_PACKAGES_DIR, exist_ok=True)
if _CUSTOM_PACKAGES_DIR not in sys.path:
    sys.path.insert(0, _CUSTOM_PACKAGES_DIR)


_HAVE_LXML = False
try:
    from lxml import html as _lxml_html
    from lxml import etree as _lxml_etree
    _HAVE_LXML = True
except ImportError:
    pass


def _strip_html(html_text: str) -> str:
    """Strip HTML to clean text. Uses lxml if available, falls back to regex."""
    if _HAVE_LXML:
        return _strip_html_lxml(html_text)
    return _strip_html_regex(html_text)


def _strip_html_lxml(html_text: str) -> str:
    """Extract clean text from HTML using lxml — much better than regex."""
    try:
        doc = _lxml_html.fromstring(html_text)
    except Exception:
        return _strip_html_regex(html_text)

    # Remove unwanted elements entirely
    for tag in ("script", "style", "svg", "noscript", "header", "footer", "nav", "iframe"):
        for el in doc.xpath(f"//{tag}"):
            el.getparent().remove(el)

    # Remove hidden elements
    for el in doc.xpath("//*[@style]"):
        style = (el.get("style") or "").lower()
        if "display:none" in style or "display: none" in style or "visibility:hidden" in style:
            el.getparent().remove(el)

    # Extract text
    text = doc.text_content()

    # Clean up
    text = _clean_text(text)
    return text


def _strip_html_regex(html_text: str) -> str:
    """Fallback regex-based HTML stripping."""
    text = _re.sub(r'<script[^>]*>.*?</script>', '', html_text, flags=_re.DOTALL | _re.IGNORECASE)
    text = _re.sub(r'<style[^>]*>.*?</style>', '', text, flags=_re.DOTALL | _re.IGNORECASE)
    text = _re.sub(r'<!--.*?-->', '', text, flags=_re.DOTALL)
    text = _re.sub(r'<img[^>]*>', '', text, flags=_re.IGNORECASE)
    text = _re.sub(r'<svg[^>]*>.*?</svg>', '', text, flags=_re.DOTALL | _re.IGNORECASE)
    text = _re.sub(r'<[^>]+>', ' ', text)
    text = _re.sub(r'!\[[^\]]*\]\([^)]*\)', '', text)
    text = _re.sub(r'https?://\S+\.(png|jpg|jpeg|gif|svg|webp|ico|bmp)\S*', '', text, flags=_re.IGNORECASE)
    text = text.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    text = text.replace('&quot;', '"').replace('&#39;', "'")
    text = _clean_text(text)
    return text


def _clean_text(text: str) -> str:
    """Shared text cleanup: normalize unicode, collapse whitespace, force ASCII."""
    text = text.replace('\u2014', '-').replace('\u2013', '-')
    text = text.replace('\u2018', "'").replace('\u2019', "'")
    text = text.replace('\u201c', '"').replace('\u201d', '"')
    text = text.replace('\u2026', '...').replace('\u00a0', ' ')
    text = text.replace('\u2022', '-').replace('\u00b7', '-')
    text = text.encode('ascii', 'replace').decode('ascii')
    text = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    text = _re.sub(r'\n\s*\n\s*\n+', '\n\n', text)
    text = _re.sub(r'[ \t]+', ' ', text)
    return text.strip()


def _extract_structured(html_text: str) -> dict:
    """Extract structured data from HTML using lxml — tables, headings, lists, links."""
    if not _HAVE_LXML:
        return {}
    try:
        doc = _lxml_html.fromstring(html_text)
    except Exception:
        return {}

    result = {}

    # Extract title
    title_el = doc.xpath("//title")
    if title_el:
        result["title"] = (title_el[0].text_content() or "").strip()

    # Extract meta description
    meta = doc.xpath("//meta[@name='description']/@content")
    if meta:
        result["description"] = meta[0]

    # Extract headings
    headings = []
    for level in range(1, 4):
        for h in doc.xpath(f"//h{level}"):
            text = (h.text_content() or "").strip()
            if text:
                headings.append(f"{'#' * level} {text}")
    if headings:
        result["headings"] = headings[:20]

    # Extract tables as CSV-style text
    tables = []
    for table in doc.xpath("//table")[:3]:  # Max 3 tables
        rows = []
        for tr in table.xpath(".//tr")[:30]:  # Max 30 rows
            cells = [(_lxml_etree.tostring(td, method="text", encoding="unicode") or "").strip()
                     for td in tr.xpath(".//td|.//th")]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            tables.append("\n".join(rows))
    if tables:
        result["tables"] = tables

    # Extract links
    links = []
    for a in doc.xpath("//a[@href]")[:20]:
        href = a.get("href", "")
        text = (a.text_content() or "").strip()
        if text and href and not href.startswith("#") and not href.startswith("javascript:"):
            links.append({"text": text[:60], "href": href})
    if links:
        result["links"] = links

    return result

from .tool_registry import registry
from .memory_manager import memory
from .skill_manager import skills
from .cron_manager import cron

# --- Sandboxed base directory ---
SANDBOX_ROOT = os.environ.get("PEGASUS_WORKSPACE_DIR", os.path.expanduser("~/Documents/pegasus_workspace"))
os.makedirs(SANDBOX_ROOT, exist_ok=True)


def _safe_path(path: str) -> str:
    """Resolve a path and ensure it's within the sandbox."""
    resolved = os.path.realpath(os.path.join(SANDBOX_ROOT, path))
    if not resolved.startswith(os.path.realpath(SANDBOX_ROOT)):
        raise PermissionError(f"Path escapes sandbox: {path}")
    return resolved


# ---- File Tools ----

def file_read(path: str) -> dict:
    """Read a file from the workspace."""
    full = _safe_path(path)
    if not os.path.isfile(full):
        return {"error": f"File not found: {path}"}
    with open(full, "r", errors="replace") as f:
        content = f.read(100_000)  # 100KB limit
    return {"path": path, "content": content}


registry.register(
    name="file_read",
    description="Read a file from the workspace. Path is relative to workspace root.",
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Relative path to file"},
        },
        "required": ["path"],
    },
    handler=file_read,
    category="filesystem",
)


def file_write(path: str, content: str) -> dict:
    """Write content to a file in the workspace."""
    full = _safe_path(path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, "w", encoding="utf-8") as f:
        f.write(content)
    actual_size = os.path.getsize(full)
    return {"path": path, "bytes_written": actual_size}


registry.register(
    name="file_write",
    description="Write content to a file in the workspace. Creates directories as needed.",
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Relative path to file"},
            "content": {"type": "string", "description": "Content to write"},
        },
        "required": ["path", "content"],
    },
    handler=file_write,
    category="filesystem",
)


def file_list(path: str = ".") -> dict:
    """List files and directories in a workspace path."""
    full = _safe_path(path)
    if not os.path.isdir(full):
        return {"error": f"Not a directory: {path}"}
    entries = []
    for name in sorted(os.listdir(full)):
        full_entry = os.path.join(full, name)
        entries.append({
            "name": name,
            "type": "dir" if os.path.isdir(full_entry) else "file",
            "size": os.path.getsize(full_entry) if os.path.isfile(full_entry) else None,
        })
    return {"path": path, "workspace_root": SANDBOX_ROOT, "entries": entries}


registry.register(
    name="file_list",
    description="List files and directories in the workspace.",
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Relative directory path (default: root)"},
        },
        "required": [],
    },
    handler=file_list,
    category="filesystem",
)


# ---- Shell Execution (emulated via Python on iOS) ----

import shutil as _shutil
import glob as _glob_mod
import fnmatch as _fnmatch
import hashlib as _hashlib
import base64 as _base64
import gzip as _gzip_mod
import tarfile as _tarfile
import difflib as _difflib


def _shell_ls(args, cwd):
    """Emulate ls."""
    show_all = False
    show_long = False
    recursive = False
    paths = []
    for a in args:
        if a.startswith("-"):
            if "a" in a: show_all = True
            if "l" in a: show_long = True
            if "R" in a: recursive = True
        else:
            paths.append(a)
    target = os.path.join(cwd, paths[0]) if paths else cwd
    if not os.path.isdir(target):
        if os.path.exists(target):
            return os.path.basename(target)
        return f"ls: {paths[0] if paths else target}: No such file or directory"

    def _list_dir(dirpath, prefix=""):
        entries = sorted(os.listdir(dirpath))
        if not show_all:
            entries = [e for e in entries if not e.startswith(".")]
        lines = []
        if prefix:
            lines.append(f"\n{prefix}:")
        if show_long:
            import time as _time
            for e in entries:
                fp = os.path.join(dirpath, e)
                try:
                    st = os.stat(fp)
                    kind = "d" if os.path.isdir(fp) else "-"
                    size = st.st_size
                    mtime = _time.strftime("%b %d %H:%M", _time.localtime(st.st_mtime))
                    lines.append(f"{kind}rw-r--r--  {size:>10}  {mtime}  {e}")
                except OSError:
                    lines.append(f"?           ?  {e}")
        else:
            lines.extend(entries)
        return lines, entries

    lines, entries = _list_dir(target)
    if recursive:
        for e in entries:
            fp = os.path.join(target, e)
            if os.path.isdir(fp):
                rel = os.path.relpath(fp, cwd)
                sub_lines, _ = _list_dir(fp, prefix=rel)
                lines.extend(sub_lines)
    return "\n".join(lines)


def _shell_cat(args, cwd):
    """Emulate cat."""
    if not args:
        return "cat: missing file"
    out = []
    for path in args:
        fp = os.path.join(cwd, path)
        if not os.path.isfile(fp):
            out.append(f"cat: {path}: No such file or directory")
            continue
        with open(fp, "r", errors="replace") as f:
            out.append(f.read(100_000))
    return "\n".join(out)


def _shell_head(args, cwd, stdin=""):
    """Emulate head."""
    n = 10
    paths = []
    i = 0
    while i < len(args):
        if args[i] == "-n" and i + 1 < len(args):
            try: n = int(args[i+1])
            except ValueError: pass
            i += 2
        elif args[i].startswith("-") and args[i][1:].isdigit():
            n = int(args[i][1:])
            i += 1
        else:
            paths.append(args[i])
            i += 1
    if not paths and stdin:
        lines = stdin.splitlines(True)
        return "".join(lines[:n])
    if not paths:
        return "head: missing file"
    fp = os.path.join(cwd, paths[0])
    if not os.path.isfile(fp):
        return f"head: {paths[0]}: No such file or directory"
    with open(fp, "r", errors="replace") as f:
        lines = []
        for _ in range(n):
            line = f.readline()
            if not line: break
            lines.append(line)
    return "".join(lines)


def _shell_tail(args, cwd, stdin=""):
    """Emulate tail."""
    n = 10
    paths = []
    i = 0
    while i < len(args):
        if args[i] == "-n" and i + 1 < len(args):
            try: n = int(args[i+1])
            except ValueError: pass
            i += 2
        elif args[i].startswith("-") and args[i][1:].isdigit():
            n = int(args[i][1:])
            i += 1
        else:
            paths.append(args[i])
            i += 1
    if not paths and stdin:
        lines = stdin.splitlines(True)
        return "".join(lines[-n:])
    if not paths:
        return "tail: missing file"
    fp = os.path.join(cwd, paths[0])
    if not os.path.isfile(fp):
        return f"tail: {paths[0]}: No such file or directory"
    with open(fp, "r", errors="replace") as f:
        all_lines = f.readlines()
    return "".join(all_lines[-n:])


def _shell_grep(args, cwd, stdin=""):
    """Emulate grep with -i, -n, -r, -c, -l, -v, -w, -o, -E, -F, --include."""
    ignore_case = False
    show_line_num = False
    recursive = False
    count_only = False
    files_only = False
    invert = False
    whole_word = False
    only_match = False
    fixed_string = False
    include_glob = None
    pattern = None
    paths = []
    i = 0
    while i < len(args):
        a = args[i]
        if a == "--include" and i + 1 < len(args):
            include_glob = args[i+1]
            i += 2
        elif a.startswith("--include="):
            include_glob = a.split("=", 1)[1].strip("'\"")
            i += 1
        elif a.startswith("-") and not a.startswith("--") and len(a) > 1:
            flags = a[1:]
            if "i" in flags: ignore_case = True
            if "n" in flags: show_line_num = True
            if "r" in flags or "R" in flags: recursive = True
            if "c" in flags: count_only = True
            if "l" in flags: files_only = True
            if "v" in flags: invert = True
            if "w" in flags: whole_word = True
            if "o" in flags: only_match = True
            if "E" in flags: pass  # extended regex is default
            if "F" in flags: fixed_string = True
            i += 1
        elif a in ("-e", "--regexp") and i + 1 < len(args):
            pattern = args[i+1]
            i += 2
        elif pattern is None:
            pattern = a
            i += 1
        else:
            paths.append(a)
            i += 1
    if pattern is None:
        return "grep: missing pattern"

    re_flags = _re.IGNORECASE if ignore_case else 0
    if fixed_string:
        pat = _re.escape(pattern)
    elif whole_word:
        pat = r'\b' + pattern + r'\b'
    else:
        pat = pattern
    try:
        regex = _re.compile(pat, re_flags)
    except _re.error as e:
        return f"grep: invalid regex: {e}"

    results = []
    file_counts = {}

    def search_lines(lines, label):
        count = 0
        for num, line in enumerate(lines, 1):
            if len(results) > 500: return
            match = regex.search(line)
            matched = bool(match) != invert  # XOR for -v
            if matched:
                count += 1
                if not count_only and not files_only:
                    prefix = f"{label}:" if label else ""
                    ln = f"{num}:" if show_line_num else ""
                    if only_match and match and not invert:
                        results.append(f"{prefix}{ln}{match.group()}")
                    else:
                        results.append(f"{prefix}{ln}{line.rstrip()}")
        if count_only:
            lbl = f"{label}:" if label else ""
            results.append(f"{lbl}{count}")
        elif files_only and count > 0:
            results.append(label)

    def search_file(fp, label):
        try:
            with open(fp, "r", errors="replace") as f:
                search_lines(f.readlines(), label)
        except (OSError, UnicodeDecodeError):
            pass

    # If no paths and we have stdin, grep stdin
    if not paths and stdin:
        search_lines(stdin.splitlines(True), "")
        return "\n".join(results) if results else ""

    if not paths:
        paths = ["."]
    multi = len(paths) > 1 or recursive

    for p in paths:
        fp = os.path.join(cwd, p)
        if os.path.isfile(fp):
            if include_glob and not _fnmatch.fnmatch(os.path.basename(fp), include_glob):
                continue
            search_file(fp, p if multi else "")
        elif os.path.isdir(fp) and recursive:
            for root, dirs, files in os.walk(fp):
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                for fname in sorted(files):
                    if include_glob and not _fnmatch.fnmatch(fname, include_glob):
                        continue
                    full = os.path.join(root, fname)
                    rel = os.path.relpath(full, cwd)
                    search_file(full, rel)
        elif os.path.isdir(fp) and not recursive:
            return f"grep: {p}: Is a directory"
    return "\n".join(results) if results else ""


def _shell_find(args, cwd):
    """Emulate find with -name, -iname, -type, -size, -mtime, -maxdepth, -exec."""
    import time as _time
    search_dir = cwd
    name_pattern = None
    iname_pattern = None
    find_type = None
    max_depth = None
    min_size = None  # in bytes
    max_size = None
    mtime_days = None  # -mtime +N or -N
    exec_cmd = None
    delete = False
    i = 0
    while i < len(args):
        a = args[i]
        if a == "-name" and i + 1 < len(args):
            name_pattern = args[i+1]
            i += 2
        elif a == "-iname" and i + 1 < len(args):
            iname_pattern = args[i+1].lower()
            i += 2
        elif a == "-type" and i + 1 < len(args):
            find_type = args[i+1]
            i += 2
        elif a == "-maxdepth" and i + 1 < len(args):
            try: max_depth = int(args[i+1])
            except ValueError: pass
            i += 2
        elif a == "-size" and i + 1 < len(args):
            sz = args[i+1]
            multiplier = 1
            if sz.endswith("k"): multiplier = 1024; sz = sz[:-1]
            elif sz.endswith("M"): multiplier = 1024*1024; sz = sz[:-1]
            elif sz.endswith("G"): multiplier = 1024*1024*1024; sz = sz[:-1]
            elif sz.endswith("c"): sz = sz[:-1]
            if sz.startswith("+"):
                try: min_size = int(sz[1:]) * multiplier
                except ValueError: pass
            elif sz.startswith("-"):
                try: max_size = int(sz[1:]) * multiplier
                except ValueError: pass
            i += 2
        elif a == "-mtime" and i + 1 < len(args):
            mtime_days = args[i+1]
            i += 2
        elif a == "-exec" :
            # Collect until ;
            exec_parts = []
            i += 1
            while i < len(args) and args[i] != ";":
                exec_parts.append(args[i])
                i += 1
            exec_cmd = " ".join(exec_parts)
            i += 1  # skip ;
        elif a == "-delete":
            delete = True
            i += 1
        elif not a.startswith("-"):
            search_dir = os.path.join(cwd, a)
            i += 1
        else:
            i += 1

    now = _time.time()
    results = []

    def matches(entry_path, entry_name, is_dir):
        if find_type == "f" and is_dir: return False
        if find_type == "d" and not is_dir: return False
        if name_pattern and not _fnmatch.fnmatch(entry_name, name_pattern): return False
        if iname_pattern and not _fnmatch.fnmatch(entry_name.lower(), iname_pattern): return False
        if not is_dir and (min_size is not None or max_size is not None):
            try:
                sz = os.path.getsize(entry_path)
                if min_size is not None and sz < min_size: return False
                if max_size is not None and sz > max_size: return False
            except OSError: return False
        if mtime_days is not None:
            try:
                mtime = os.path.getmtime(entry_path)
                age_days = (now - mtime) / 86400
                if mtime_days.startswith("+"):
                    if age_days <= int(mtime_days[1:]): return False
                elif mtime_days.startswith("-"):
                    if age_days >= int(mtime_days[1:]): return False
                else:
                    if int(age_days) != int(mtime_days): return False
            except (OSError, ValueError): return False
        return True

    for root, dirs, files in os.walk(search_dir):
        if max_depth is not None:
            depth = root.replace(search_dir, "").count(os.sep)
            if depth >= max_depth:
                dirs.clear()
        for d in list(dirs):
            fp = os.path.join(root, d)
            rel = os.path.relpath(fp, cwd)
            if matches(fp, d, True):
                results.append(rel)
        for f in files:
            fp = os.path.join(root, f)
            rel = os.path.relpath(fp, cwd)
            if matches(fp, f, False):
                results.append(rel)
        if len(results) > 500: break

    if delete:
        deleted = 0
        for r in results:
            fp = os.path.join(cwd, r)
            try:
                if os.path.isdir(fp):
                    _shutil.rmtree(fp)
                else:
                    os.remove(fp)
                deleted += 1
            except OSError:
                pass
        return f"Deleted {deleted} items"

    if exec_cmd:
        outputs = []
        for r in results[:100]:
            cmd = exec_cmd.replace("{}", r)
            out = _exec_single_command(cmd, cwd)
            if out:
                outputs.append(out)
        return "\n".join(outputs)

    return "\n".join(sorted(results)) if results else ""


def _shell_wc(args, cwd, stdin=""):
    """Emulate wc with -l, -w, -c flags."""
    only_lines = "-l" in args
    only_words = "-w" in args
    only_chars = "-c" in args or "-m" in args
    paths = [a for a in args if not a.startswith("-")]

    def count_text(content, label=""):
        lines = content.count("\n")
        words = len(content.split())
        chars = len(content)
        if only_lines: return f"  {lines} {label}".rstrip()
        if only_words: return f"  {words} {label}".rstrip()
        if only_chars: return f"  {chars} {label}".rstrip()
        return f"  {lines}  {words}  {chars} {label}".rstrip()

    if not paths and stdin:
        return count_text(stdin)
    if not paths:
        return "wc: missing file"
    out = []
    for p in paths:
        fp = os.path.join(cwd, p)
        if not os.path.isfile(fp):
            out.append(f"wc: {p}: No such file or directory")
            continue
        with open(fp, "r", errors="replace") as f:
            content = f.read()
        out.append(count_text(content, p))
    return "\n".join(out)


def _shell_mkdir(args, cwd):
    """Emulate mkdir."""
    parents = "-p" in args
    paths = [a for a in args if not a.startswith("-")]
    if not paths:
        return "mkdir: missing directory name"
    for p in paths:
        fp = os.path.join(cwd, p)
        if parents:
            os.makedirs(fp, exist_ok=True)
        else:
            os.makedirs(fp, exist_ok=False)
    return ""


def _shell_rm(args, cwd):
    """Emulate rm."""
    recursive = False
    force = False
    paths = []
    for a in args:
        if a.startswith("-"):
            if "r" in a or "R" in a: recursive = True
            if "f" in a: force = True
        else:
            paths.append(a)
    if not paths:
        return "rm: missing operand"
    for p in paths:
        fp = os.path.join(cwd, p)
        if os.path.isdir(fp):
            if recursive:
                _shutil.rmtree(fp)
            else:
                return f"rm: {p}: is a directory (use -r)"
        elif os.path.exists(fp):
            os.remove(fp)
        elif not force:
            return f"rm: {p}: No such file or directory"
    return ""


def _shell_cp(args, cwd):
    """Emulate cp."""
    recursive = "-r" in args or "-R" in args
    paths = [a for a in args if not a.startswith("-")]
    if len(paths) < 2:
        return "cp: missing destination"
    src = os.path.join(cwd, paths[0])
    dst = os.path.join(cwd, paths[1])
    if os.path.isdir(src):
        if recursive:
            _shutil.copytree(src, dst, dirs_exist_ok=True)
        else:
            return f"cp: {paths[0]}: is a directory (use -r)"
    else:
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        _shutil.copy2(src, dst)
    return ""


def _shell_mv(args, cwd):
    """Emulate mv."""
    paths = [a for a in args if not a.startswith("-")]
    if len(paths) < 2:
        return "mv: missing destination"
    src = os.path.join(cwd, paths[0])
    dst = os.path.join(cwd, paths[1])
    _shutil.move(src, dst)
    return ""


def _shell_echo(args, cwd):
    """Emulate echo with -n and -e."""
    no_newline = False
    interpret_escapes = False
    text_args = []
    for a in args:
        if a == "-n": no_newline = True
        elif a == "-e": interpret_escapes = True
        elif a == "-en" or a == "-ne": no_newline = True; interpret_escapes = True
        else: text_args.append(a)
    text = " ".join(text_args)
    if interpret_escapes:
        text = text.replace("\\n", "\n").replace("\\t", "\t").replace("\\\\", "\\")
    return text


def _shell_pwd(args, cwd):
    """Emulate pwd."""
    return cwd


def _shell_curl(args, cwd):
    """Emulate curl with GET/POST, -H, -d, -o, -X."""
    url = None
    method = "GET"
    headers = {"User-Agent": "Pegasus/1.0"}
    data = None
    output_file = None
    i = 0
    while i < len(args):
        a = args[i]
        if a in ("-L", "-s", "-S", "-k", "--insecure", "--silent", "--compressed", "-f", "--fail"):
            i += 1
        elif a in ("-X", "--request") and i + 1 < len(args):
            method = args[i+1]
            i += 2
        elif a in ("-H", "--header") and i + 1 < len(args):
            hdr = args[i+1]
            if ":" in hdr:
                k, v = hdr.split(":", 1)
                headers[k.strip()] = v.strip()
            i += 2
        elif a in ("-d", "--data", "--data-raw") and i + 1 < len(args):
            data = args[i+1].encode("utf-8")
            if method == "GET": method = "POST"
            i += 2
        elif a in ("-o", "--output") and i + 1 < len(args):
            output_file = args[i+1]
            i += 2
        elif not a.startswith("-"):
            url = a
            i += 1
        else:
            i += 1
    if not url:
        return "curl: missing URL"
    try:
        req = urllib.request.Request(url, headers=headers, method=method, data=data)
        with urllib.request.urlopen(req, timeout=30, context=_ssl_ctx) as resp:
            raw = resp.read(200_000)
        if output_file:
            fp = os.path.join(cwd, output_file)
            with open(fp, "wb") as f:
                f.write(raw)
            return f"Saved to {output_file} ({len(raw)} bytes)"
        return raw.decode("utf-8", errors="replace")[:50_000]
    except Exception as e:
        return f"curl: {e}"


def _shell_wget(args, cwd):
    """Emulate wget (download file)."""
    url = None
    output = None
    i = 0
    while i < len(args):
        if args[i] in ("-O", "-o", "--output-document") and i + 1 < len(args):
            output = args[i+1]
            i += 2
        elif not args[i].startswith("-"):
            url = args[i]
            i += 1
        else:
            i += 1
    if not url:
        return "wget: missing URL"
    if not output:
        output = url.split("/")[-1].split("?")[0] or "download"
    fp = os.path.join(cwd, output)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Pegasus/1.0"})
        with urllib.request.urlopen(req, timeout=60, context=_ssl_ctx) as resp:
            data = resp.read(50_000_000)  # 50MB max
        os.makedirs(os.path.dirname(fp) or cwd, exist_ok=True)
        with open(fp, "wb") as f:
            f.write(data)
        return f"Saved {output} ({len(data)} bytes)"
    except Exception as e:
        return f"wget: {e}"


def _shell_touch(args, cwd):
    """Emulate touch."""
    for a in args:
        if a.startswith("-"): continue
        fp = os.path.join(cwd, a)
        os.makedirs(os.path.dirname(fp) or cwd, exist_ok=True)
        Path(fp).touch()
    return ""


def _shell_du(args, cwd):
    """Emulate du -sh."""
    paths = [a for a in args if not a.startswith("-")]
    target = os.path.join(cwd, paths[0]) if paths else cwd
    total = 0
    for root, dirs, files in os.walk(target):
        for f in files:
            try:
                total += os.path.getsize(os.path.join(root, f))
            except OSError:
                pass
    if total < 1024:
        size_str = f"{total}B"
    elif total < 1024 * 1024:
        size_str = f"{total/1024:.1f}K"
    elif total < 1024 * 1024 * 1024:
        size_str = f"{total/(1024*1024):.1f}M"
    else:
        size_str = f"{total/(1024*1024*1024):.1f}G"
    return f"{size_str}\t{paths[0] if paths else '.'}"


def _shell_which(args, cwd):
    """Emulate which - report available commands."""
    extra = {"python", "python3", "pip", "pip3"}
    out = []
    for a in args:
        if a.startswith("-"): continue
        if a in _SHELL_COMMANDS or a in extra:
            out.append(f"/usr/bin/{a} (emulated via Python)")
        elif a in _CLI_ADAPTERS:
            out.append(f"/usr/bin/{a} (CLI adapter)")
        else:
            out.append(f"{a} not found")
    if not args:
        # List all available commands
        all_cmds = sorted(set(list(_SHELL_COMMANDS.keys()) + list(extra) + list(_CLI_ADAPTERS.keys())))
        return "Available commands:\n" + ", ".join(all_cmds) + "\n\nPip-installed packages with CLI entry points are also auto-discovered."
    return "\n".join(out)


def _shell_env(args, cwd):
    """Emulate env."""
    return "\n".join(f"{k}={v}" for k, v in sorted(os.environ.items()))


def _shell_date(args, cwd):
    """Emulate date."""
    import datetime
    return datetime.datetime.now().strftime("%a %b %d %H:%M:%S %Z %Y")


def _shell_sort(args, cwd, stdin=""):
    """Emulate sort with -r, -n, -u, -k, -t."""
    reverse = "-r" in args
    numeric = "-n" in args
    unique = "-u" in args
    key_field = None
    separator = None
    paths = []
    i = 0
    while i < len(args):
        if args[i] == "-k" and i + 1 < len(args):
            try: key_field = int(args[i+1].split(",")[0]) - 1
            except ValueError: pass
            i += 2
        elif args[i] == "-t" and i + 1 < len(args):
            separator = args[i+1]
            i += 2
        elif not args[i].startswith("-"):
            paths.append(args[i])
            i += 1
        else:
            i += 1

    if paths:
        fp = os.path.join(cwd, paths[0])
        with open(fp, "r", errors="replace") as f:
            lines = f.readlines()
    elif stdin:
        lines = stdin.splitlines(True)
    else:
        return ""

    def sort_key(line):
        if key_field is not None:
            parts = line.split(separator) if separator else line.split()
            val = parts[key_field].strip() if key_field < len(parts) else ""
        else:
            val = line.strip()
        if numeric:
            try: return float(_re.match(r'-?[\d.]+', val).group())
            except (AttributeError, ValueError): return 0
        return val

    lines.sort(key=sort_key, reverse=reverse)
    if unique:
        seen = set()
        deduped = []
        for line in lines:
            k = sort_key(line)
            if k not in seen:
                seen.add(k)
                deduped.append(line)
        lines = deduped
    return "".join(lines[:5000])


def _shell_uniq(args, cwd, stdin=""):
    """Emulate uniq with -c, -d, -u."""
    count_mode = "-c" in args
    dupes_only = "-d" in args
    unique_only = "-u" in args
    paths = [a for a in args if not a.startswith("-")]

    if paths:
        fp = os.path.join(cwd, paths[0])
        with open(fp, "r", errors="replace") as f:
            lines = f.readlines()
    elif stdin:
        lines = stdin.splitlines(True)
    else:
        return ""

    result = []
    prev = None
    count = 0
    for line in lines:
        if line == prev:
            count += 1
        else:
            if prev is not None:
                include = True
                if dupes_only and count < 2: include = False
                if unique_only and count > 1: include = False
                if include:
                    if count_mode:
                        result.append(f"  {count} {prev.rstrip()}")
                    else:
                        result.append(prev.rstrip())
            prev = line
            count = 1
    if prev is not None:
        include = True
        if dupes_only and count < 2: include = False
        if unique_only and count > 1: include = False
        if include:
            if count_mode:
                result.append(f"  {count} {prev.rstrip()}")
            else:
                result.append(prev.rstrip())
    return "\n".join(result[:5000])


def _shell_sed(args, cwd, stdin=""):
    """Emulate sed with regex support, -i, -n, multiple expressions."""
    exprs = []
    in_place = False
    quiet = False
    paths = []
    i = 0
    while i < len(args):
        a = args[i]
        if a == "-i":
            in_place = True
            i += 1
        elif a == "-n":
            quiet = True
            i += 1
        elif a == "-e" and i + 1 < len(args):
            exprs.append(args[i+1])
            i += 2
        elif (a.startswith("s") and "/" in a) or (a.startswith("'") and "s" in a):
            exprs.append(a.strip("'\""))
            i += 1
        elif not a.startswith("-"):
            if not exprs:
                exprs.append(a.strip("'\""))
            else:
                paths.append(a)
            i += 1
        else:
            i += 1

    if not exprs:
        return "sed: missing expression"

    def apply_sed(content):
        for expr in exprs:
            if not expr.startswith("s"):
                continue
            delim = expr[1]
            parts = expr[2:].split(delim)
            if len(parts) < 2:
                continue
            old_pat = parts[0]
            new_pat = parts[1]
            flags_str = parts[2] if len(parts) > 2 else ""
            re_flags = 0
            if "i" in flags_str: re_flags |= _re.IGNORECASE
            count = 0 if "g" in flags_str else 1
            try:
                content = _re.sub(old_pat, new_pat, content, count=count, flags=re_flags)
            except _re.error:
                # Fallback to literal replace
                if "g" in flags_str:
                    content = content.replace(old_pat, new_pat)
                else:
                    content = content.replace(old_pat, new_pat, 1)
        return content

    if paths:
        fp = os.path.join(cwd, paths[0])
        if not os.path.isfile(fp):
            return f"sed: {paths[0]}: No such file"
        with open(fp, "r", errors="replace") as f:
            content = f.read()
        result = apply_sed(content)
        if in_place:
            with open(fp, "w", encoding="utf-8") as f:
                f.write(result)
            return ""
        return result[:50_000]
    elif stdin:
        return apply_sed(stdin)
    return "sed: missing input file"


def _shell_awk(args, cwd, stdin=""):
    """Emulate basic awk: field extraction, patterns, BEGIN/END, -F."""
    separator = None
    program = None
    paths = []
    i = 0
    while i < len(args):
        if args[i] == "-F" and i + 1 < len(args):
            separator = args[i+1]
            i += 2
        elif program is None and (args[i].startswith("{") or args[i].startswith("'") or
                                   args[i].startswith("/")):
            program = args[i].strip("'\"")
            i += 1
        elif program is None:
            program = args[i].strip("'\"")
            i += 1
        else:
            paths.append(args[i])
            i += 1

    if not program:
        return "awk: missing program"

    if paths:
        fp = os.path.join(cwd, paths[0])
        with open(fp, "r", errors="replace") as f:
            lines = f.readlines()
    elif stdin:
        lines = stdin.splitlines(True)
    else:
        return ""

    # Simple awk: handle {print $N} and /pattern/{print $N}
    output = []
    pat_match = _re.match(r'/([^/]+)/\s*\{(.+)\}', program)
    simple_match = _re.match(r'\{(.+)\}', program)

    if pat_match:
        pattern = pat_match.group(1)
        action = pat_match.group(2).strip()
    elif simple_match:
        pattern = None
        action = simple_match.group(1).strip()
    else:
        pattern = None
        action = f"print {program}" if "$" in program else program

    for line_num, line in enumerate(lines, 1):
        line = line.rstrip("\n")
        if pattern and not _re.search(pattern, line):
            continue
        fields = line.split(separator) if separator else line.split()
        fields = [""] + fields  # $0=whole line at conceptual level

        if action.startswith("print"):
            print_expr = action[5:].strip()
            if not print_expr:
                output.append(line)
            else:
                parts = []
                for token in _re.split(r'[,\s]+', print_expr):
                    token = token.strip('"')
                    if token.startswith("$"):
                        try:
                            idx = int(token[1:])
                            if idx == 0:
                                parts.append(line)
                            elif idx <= len(fields) - 1:
                                parts.append(fields[idx])
                            else:
                                parts.append("")
                        except ValueError:
                            parts.append(token)
                    elif token == "NR":
                        parts.append(str(line_num))
                    elif token == "NF":
                        parts.append(str(len(fields) - 1))
                    else:
                        parts.append(token)
                output.append(" ".join(parts))
        if len(output) > 5000: break
    return "\n".join(output)


def _shell_cut(args, cwd, stdin=""):
    """Emulate cut -d -f."""
    delimiter = "\t"
    fields_spec = None
    char_spec = None
    paths = []
    i = 0
    while i < len(args):
        if args[i] == "-d" and i + 1 < len(args):
            delimiter = args[i+1]
            i += 2
        elif args[i].startswith("-d") and len(args[i]) > 2:
            delimiter = args[i][2:]
            i += 1
        elif args[i] == "-f" and i + 1 < len(args):
            fields_spec = args[i+1]
            i += 2
        elif args[i].startswith("-f") and len(args[i]) > 2:
            fields_spec = args[i][2:]
            i += 1
        elif args[i] == "-c" and i + 1 < len(args):
            char_spec = args[i+1]
            i += 2
        elif not args[i].startswith("-"):
            paths.append(args[i])
            i += 1
        else:
            i += 1

    if paths:
        fp = os.path.join(cwd, paths[0])
        with open(fp, "r", errors="replace") as f:
            lines = f.readlines()
    elif stdin:
        lines = stdin.splitlines(True)
    else:
        return ""

    def parse_spec(spec):
        """Parse field spec like '1,3' or '1-3' or '2-'."""
        indices = set()
        for part in spec.split(","):
            if "-" in part:
                a, b = part.split("-", 1)
                start = int(a) if a else 1
                end = int(b) if b else 999
                indices.update(range(start, end + 1))
            else:
                indices.add(int(part))
        return sorted(indices)

    output = []
    for line in lines:
        line = line.rstrip("\n")
        if char_spec:
            indices = parse_spec(char_spec)
            output.append("".join(line[i-1] for i in indices if i-1 < len(line)))
        elif fields_spec:
            parts = line.split(delimiter)
            indices = parse_spec(fields_spec)
            selected = [parts[i-1] for i in indices if i-1 < len(parts)]
            output.append(delimiter.join(selected))
        else:
            output.append(line)
    return "\n".join(output[:5000])


def _shell_tr(args, cwd, stdin=""):
    """Emulate tr for character translation/deletion."""
    delete = "-d" in args
    squeeze = "-s" in args
    sets = [a for a in args if not a.startswith("-")]
    if not stdin:
        return "tr: requires piped input"
    if delete and sets:
        chars_to_delete = sets[0].strip("'\"")
        result = stdin
        for c in chars_to_delete:
            result = result.replace(c, "")
        return result
    if len(sets) >= 2:
        set1 = sets[0].strip("'\"")
        set2 = sets[1].strip("'\"")
        # Handle character classes
        if set1 == "[:upper:]": set1 = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        if set1 == "[:lower:]": set1 = "abcdefghijklmnopqrstuvwxyz"
        if set2 == "[:upper:]": set2 = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        if set2 == "[:lower:]": set2 = "abcdefghijklmnopqrstuvwxyz"
        table = str.maketrans(set1[:len(set2)], set2[:len(set1)])
        result = stdin.translate(table)
        if squeeze:
            for c in set2:
                while c + c in result:
                    result = result.replace(c + c, c)
        return result
    return stdin


def _shell_tee(args, cwd, stdin=""):
    """Emulate tee."""
    append = "-a" in args
    paths = [a for a in args if not a.startswith("-")]
    for p in paths:
        fp = os.path.join(cwd, p)
        mode = "a" if append else "w"
        with open(fp, mode, encoding="utf-8") as f:
            f.write(stdin)
    return stdin


def _shell_xargs(args, cwd, stdin=""):
    """Emulate xargs."""
    if not args:
        return stdin
    cmd_template = " ".join(args)
    items = stdin.split()
    outputs = []
    for item in items[:100]:
        if "{}" in cmd_template:
            cmd = cmd_template.replace("{}", item)
        else:
            cmd = f"{cmd_template} {item}"
        out = _exec_single_command(cmd, cwd)
        if out:
            outputs.append(out)
    return "\n".join(outputs)


def _shell_diff(args, cwd):
    """Emulate diff."""
    unified = "-u" in args
    paths = [a for a in args if not a.startswith("-")]
    if len(paths) < 2:
        return "diff: need two files"
    fp1 = os.path.join(cwd, paths[0])
    fp2 = os.path.join(cwd, paths[1])
    if not os.path.isfile(fp1): return f"diff: {paths[0]}: No such file"
    if not os.path.isfile(fp2): return f"diff: {paths[1]}: No such file"
    with open(fp1, "r", errors="replace") as f:
        lines1 = f.readlines()
    with open(fp2, "r", errors="replace") as f:
        lines2 = f.readlines()
    if unified:
        result = _difflib.unified_diff(lines1, lines2, fromfile=paths[0], tofile=paths[1])
    else:
        result = _difflib.ndiff(lines1, lines2)
    return "".join(list(result)[:5000])


def _shell_tar(args, cwd):
    """Emulate tar."""
    create = False
    extract = False
    listt = False
    gzipped = False
    filename = None
    paths = []
    i = 0
    while i < len(args):
        a = args[i]
        if a.startswith("-") or (i == 0 and not os.path.exists(os.path.join(cwd, a))):
            flags = a.lstrip("-")
            if "c" in flags: create = True
            if "x" in flags: extract = True
            if "t" in flags: listt = True
            if "z" in flags: gzipped = True
            if "f" in flags:
                if i + 1 < len(args):
                    filename = args[i+1]
                    i += 2
                    continue
            i += 1
        else:
            paths.append(a)
            i += 1

    if not filename:
        return "tar: missing -f filename"

    fp = os.path.join(cwd, filename)

    if create:
        mode = "w:gz" if gzipped else "w"
        with _tarfile.open(fp, mode) as tar:
            for p in paths:
                src = os.path.join(cwd, p)
                tar.add(src, arcname=p)
        return f"Created {filename}"

    if extract:
        mode = "r:gz" if gzipped or filename.endswith(".gz") or filename.endswith(".tgz") else "r"
        with _tarfile.open(fp, mode) as tar:
            tar.extractall(path=cwd)
        return f"Extracted {filename}"

    if listt:
        mode = "r:gz" if gzipped or filename.endswith(".gz") or filename.endswith(".tgz") else "r"
        with _tarfile.open(fp, mode) as tar:
            return "\n".join(tar.getnames()[:500])

    return "tar: specify -c, -x, or -t"


def _shell_zip(args, cwd):
    """Emulate zip."""
    paths = [a for a in args if not a.startswith("-")]
    if len(paths) < 2:
        return "zip: need zipfile and files"
    zipname = paths[0]
    if not zipname.endswith(".zip"): zipname += ".zip"
    fp = os.path.join(cwd, zipname)
    with zipfile.ZipFile(fp, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in paths[1:]:
            src = os.path.join(cwd, p)
            if os.path.isdir(src):
                for root, dirs, files in os.walk(src):
                    for f in files:
                        full = os.path.join(root, f)
                        arcname = os.path.relpath(full, cwd)
                        zf.write(full, arcname)
            elif os.path.isfile(src):
                zf.write(src, p)
    return f"Created {zipname}"


def _shell_unzip(args, cwd):
    """Emulate unzip."""
    listt = "-l" in args
    paths = [a for a in args if not a.startswith("-")]
    if not paths:
        return "unzip: missing file"
    fp = os.path.join(cwd, paths[0])
    if not os.path.isfile(fp):
        return f"unzip: {paths[0]}: No such file"
    with zipfile.ZipFile(fp, "r") as zf:
        if listt:
            lines = []
            for info in zf.infolist():
                lines.append(f"  {info.file_size:>10}  {info.filename}")
            return "\n".join(lines[:500])
        zf.extractall(cwd)
    return f"Extracted {paths[0]}"


def _shell_gzip(args, cwd):
    """Emulate gzip/gunzip."""
    decompress = "-d" in args
    keep = "-k" in args
    paths = [a for a in args if not a.startswith("-")]
    if not paths: return "gzip: missing file"
    for p in paths:
        fp = os.path.join(cwd, p)
        if decompress or p.endswith(".gz"):
            out = fp[:-3] if fp.endswith(".gz") else fp + ".out"
            with _gzip_mod.open(fp, "rb") as f_in:
                with open(out, "wb") as f_out:
                    f_out.write(f_in.read())
            if not keep: os.remove(fp)
            return f"Decompressed to {os.path.basename(out)}"
        else:
            out = fp + ".gz"
            with open(fp, "rb") as f_in:
                with _gzip_mod.open(out, "wb") as f_out:
                    f_out.write(f_in.read())
            if not keep: os.remove(fp)
            return f"Compressed to {os.path.basename(out)}"


def _shell_basename(args, cwd):
    return os.path.basename(args[0]) if args else ""

def _shell_dirname(args, cwd):
    return os.path.dirname(args[0]) if args else ""

def _shell_realpath(args, cwd):
    return os.path.realpath(os.path.join(cwd, args[0])) if args else cwd

def _shell_readlink(args, cwd):
    fp = os.path.join(cwd, args[0]) if args else ""
    try: return os.readlink(fp)
    except OSError: return fp


def _shell_ln(args, cwd):
    """Emulate ln [-s]."""
    symbolic = "-s" in args
    paths = [a for a in args if not a.startswith("-")]
    if len(paths) < 2: return "ln: need target and link name"
    target = os.path.join(cwd, paths[0])
    link = os.path.join(cwd, paths[1])
    if symbolic:
        os.symlink(target, link)
    else:
        os.link(target, link)
    return ""


def _shell_stat(args, cwd):
    """Emulate stat."""
    import time as _time
    paths = [a for a in args if not a.startswith("-")]
    if not paths: return "stat: missing file"
    fp = os.path.join(cwd, paths[0])
    try:
        st = os.stat(fp)
        kind = "directory" if os.path.isdir(fp) else ("symlink" if os.path.islink(fp) else "regular file")
        return (f"  File: {paths[0]}\n"
                f"  Size: {st.st_size}\tType: {kind}\n"
                f"  Mode: {oct(st.st_mode)}\n"
                f"  Modify: {_time.ctime(st.st_mtime)}\n"
                f"  Access: {_time.ctime(st.st_atime)}\n"
                f"  Change: {_time.ctime(st.st_ctime)}")
    except OSError as e:
        return f"stat: {paths[0]}: {e}"


def _shell_file(args, cwd):
    """Emulate file command (basic type detection)."""
    paths = [a for a in args if not a.startswith("-")]
    if not paths: return "file: missing file"
    out = []
    for p in paths:
        fp = os.path.join(cwd, p)
        if not os.path.exists(fp):
            out.append(f"{p}: No such file"); continue
        if os.path.isdir(fp):
            out.append(f"{p}: directory"); continue
        with open(fp, "rb") as f:
            header = f.read(16)
        if header[:4] == b'\x89PNG': ftype = "PNG image"
        elif header[:2] == b'\xff\xd8': ftype = "JPEG image"
        elif header[:4] == b'GIF8': ftype = "GIF image"
        elif header[:4] == b'%PDF': ftype = "PDF document"
        elif header[:2] == b'PK': ftype = "Zip archive"
        elif header[:3] == b'\x1f\x8b\x08': ftype = "gzip compressed"
        elif header[:5] == b'{\\rtf': ftype = "Rich Text Format"
        elif header[:4] in (b'\xfe\xff', b'\xff\xfe'): ftype = "Unicode text"
        else:
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    f.read(1024)
                ftype = "ASCII/UTF-8 text"
            except (UnicodeDecodeError, ValueError):
                ftype = "binary data"
        sz = os.path.getsize(fp)
        out.append(f"{p}: {ftype}, {sz} bytes")
    return "\n".join(out)


def _shell_md5sum(args, cwd):
    paths = [a for a in args if not a.startswith("-")]
    out = []
    for p in paths:
        fp = os.path.join(cwd, p)
        if not os.path.isfile(fp): out.append(f"md5sum: {p}: No such file"); continue
        h = _hashlib.md5()
        with open(fp, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""): h.update(chunk)
        out.append(f"{h.hexdigest()}  {p}")
    return "\n".join(out)


def _shell_sha256sum(args, cwd):
    paths = [a for a in args if not a.startswith("-")]
    out = []
    for p in paths:
        fp = os.path.join(cwd, p)
        if not os.path.isfile(fp): out.append(f"sha256sum: {p}: No such file"); continue
        h = _hashlib.sha256()
        with open(fp, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""): h.update(chunk)
        out.append(f"{h.hexdigest()}  {p}")
    return "\n".join(out)


def _shell_base64_cmd(args, cwd, stdin=""):
    """Emulate base64 encode/decode."""
    decode = "-d" in args or "--decode" in args
    paths = [a for a in args if not a.startswith("-")]
    if paths:
        fp = os.path.join(cwd, paths[0])
        with open(fp, "rb") as f: data = f.read()
    elif stdin:
        data = stdin.encode("utf-8")
    else:
        return "base64: missing input"
    if decode:
        return _base64.b64decode(data).decode("utf-8", errors="replace")
    return _base64.b64encode(data).decode("ascii")


def _shell_xxd(args, cwd, stdin=""):
    """Emulate xxd (hex dump)."""
    paths = [a for a in args if not a.startswith("-")]
    if paths:
        fp = os.path.join(cwd, paths[0])
        with open(fp, "rb") as f: data = f.read(4096)
    elif stdin:
        data = stdin.encode("utf-8")[:4096]
    else:
        return "xxd: missing input"
    lines = []
    for i in range(0, len(data), 16):
        chunk = data[i:i+16]
        hex_part = " ".join(f"{b:02x}" for b in chunk)
        ascii_part = "".join(chr(b) if 32 <= b < 127 else "." for b in chunk)
        lines.append(f"{i:08x}: {hex_part:<48s}  {ascii_part}")
    return "\n".join(lines)


def _shell_seq(args, cwd):
    start, step, end = 1, 1, 1
    nums = [a for a in args if not a.startswith("-") or a.lstrip("-").replace(".", "").isdigit()]
    if len(nums) == 1: end = int(nums[0])
    elif len(nums) == 2: start = int(nums[0]); end = int(nums[1])
    elif len(nums) >= 3: start = int(nums[0]); step = int(nums[1]); end = int(nums[2])
    return "\n".join(str(i) for i in range(start, end + 1, step))[:50_000]


def _shell_sleep(args, cwd):
    import time as _time
    if args:
        try:
            secs = float(args[0].rstrip("s"))
            _time.sleep(min(secs, 30))  # Cap at 30s
        except ValueError: pass
    return ""


def _shell_tac(args, cwd, stdin=""):
    if args:
        fp = os.path.join(cwd, args[0])
        with open(fp, "r", errors="replace") as f: lines = f.readlines()
    elif stdin:
        lines = stdin.splitlines(True)
    else:
        return ""
    return "".join(reversed(lines))


def _shell_rev(args, cwd, stdin=""):
    if args:
        fp = os.path.join(cwd, args[0])
        with open(fp, "r", errors="replace") as f: lines = f.readlines()
    elif stdin:
        lines = stdin.splitlines(True)
    else:
        return ""
    return "\n".join(line.rstrip()[::-1] for line in lines)


def _shell_nl(args, cwd, stdin=""):
    if args and not args[0].startswith("-"):
        fp = os.path.join(cwd, args[0])
        with open(fp, "r", errors="replace") as f: lines = f.readlines()
    elif stdin:
        lines = stdin.splitlines(True)
    else:
        return ""
    return "\n".join(f"  {i:>4}\t{line.rstrip()}" for i, line in enumerate(lines, 1))


def _shell_paste(args, cwd):
    delimiter = "\t"
    paths = []
    i = 0
    while i < len(args):
        if args[i] == "-d" and i + 1 < len(args):
            delimiter = args[i+1]
            i += 2
        elif not args[i].startswith("-"):
            paths.append(args[i])
            i += 1
        else:
            i += 1
    if len(paths) < 2: return "paste: need at least two files"
    file_lines = []
    for p in paths:
        fp = os.path.join(cwd, p)
        with open(fp, "r", errors="replace") as f:
            file_lines.append(f.readlines())
    max_lines = max(len(fl) for fl in file_lines)
    out = []
    for i in range(max_lines):
        parts = [fl[i].rstrip("\n") if i < len(fl) else "" for fl in file_lines]
        out.append(delimiter.join(parts))
    return "\n".join(out[:5000])


def _shell_column(args, cwd, stdin=""):
    """Emulate column -t."""
    table_mode = "-t" in args
    separator = None
    i = 0
    while i < len(args):
        if args[i] == "-s" and i + 1 < len(args):
            separator = args[i+1]; i += 2
        else:
            i += 1
    if not stdin: return ""
    lines = stdin.splitlines()
    if table_mode:
        rows = [line.split(separator) if separator else line.split() for line in lines]
        if not rows: return ""
        max_cols = max(len(r) for r in rows)
        widths = [0] * max_cols
        for row in rows:
            for ci, cell in enumerate(row):
                widths[ci] = max(widths[ci], len(cell))
        out = []
        for row in rows:
            parts = [cell.ljust(widths[ci]) for ci, cell in enumerate(row)]
            out.append("  ".join(parts).rstrip())
        return "\n".join(out)
    return stdin


def _shell_cal(args, cwd):
    import calendar
    import datetime
    now = datetime.datetime.now()
    if len(args) >= 2:
        try: return calendar.month(int(args[1]), int(args[0]))
        except (ValueError, IndexError): pass
    if len(args) == 1:
        try: return calendar.calendar(int(args[0]))
        except ValueError: pass
    return calendar.month(now.year, now.month)


def _shell_expr(args, cwd):
    """Emulate expr for basic arithmetic."""
    expr_str = " ".join(args)
    try:
        # Only allow safe arithmetic
        allowed = set("0123456789+-*/%() .")
        if all(c in allowed for c in expr_str):
            return str(eval(expr_str))  # safe: only digits and operators
    except Exception:
        pass
    return "0"


def _shell_test(args, cwd):
    """Emulate test / [ command."""
    if not args: return "1"
    # Remove trailing ] if present
    if args[-1] == "]": args = args[:-1]
    if len(args) == 1:
        return "0" if args[0] else "1"
    if len(args) == 2:
        op, path = args[0], os.path.join(cwd, args[1])
        if op == "-f": return "0" if os.path.isfile(path) else "1"
        if op == "-d": return "0" if os.path.isdir(path) else "1"
        if op == "-e": return "0" if os.path.exists(path) else "1"
        if op == "-s": return "0" if os.path.exists(path) and os.path.getsize(path) > 0 else "1"
        if op == "-r" or op == "-w": return "0" if os.path.exists(path) else "1"
        if op == "-z": return "0" if not args[1] else "1"
        if op == "-n": return "0" if args[1] else "1"
    if len(args) == 3:
        a, op, b = args
        if op == "=": return "0" if a == b else "1"
        if op == "!=": return "0" if a != b else "1"
        try:
            ai, bi = int(a), int(b)
            if op == "-eq": return "0" if ai == bi else "1"
            if op == "-ne": return "0" if ai != bi else "1"
            if op == "-lt": return "0" if ai < bi else "1"
            if op == "-gt": return "0" if ai > bi else "1"
            if op == "-le": return "0" if ai <= bi else "1"
            if op == "-ge": return "0" if ai >= bi else "1"
        except ValueError: pass
    return "1"


def _shell_export(args, cwd):
    for a in args:
        if "=" in a:
            k, v = a.split("=", 1)
            os.environ[k] = v
        elif a in os.environ:
            pass  # already exported
    return ""


def _shell_cat_n(args, cwd):
    """cat with -n (number lines)."""
    number = "-n" in args
    paths = [a for a in args if not a.startswith("-")]
    if not paths: return "cat: missing file"
    out = []
    for path in paths:
        fp = os.path.join(cwd, path)
        if not os.path.isfile(fp):
            out.append(f"cat: {path}: No such file or directory"); continue
        with open(fp, "r", errors="replace") as f:
            lines = f.readlines()
        if number:
            for i, line in enumerate(lines, 1):
                out.append(f"  {i:>4}\t{line.rstrip()}")
        else:
            out.append("".join(lines[:100_000]))
    return "\n".join(out)


def _shell_chmod(args, cwd):
    return "(chmod: no-op on iOS)"

def _shell_chown(args, cwd):
    return "(chown: no-op on iOS)"

def _shell_whoami(args, cwd):
    return "mobile"

def _shell_hostname(args, cwd):
    import platform
    return platform.node() or "pegasus-ios"

def _shell_uname(args, cwd):
    import platform
    flags = "".join(a for a in args if a.startswith("-"))
    if "-a" in flags:
        return f"{platform.system()} {platform.node()} {platform.release()} {platform.version()} {platform.machine()}"
    return f"{platform.system()} {platform.release()} {platform.machine()}"


# ---- CLI Adapter System ----
# Routes CLI tool invocations to their Python library equivalents.
# Since subprocess doesn't work on iOS, this lets the agent run tools like
# jq, sqlite3, yt-dlp, black, etc. by calling their Python APIs directly.

def _cli_run_module(module_name, args, cwd):
    """Run a Python module's CLI entry point (like `python -m module`)."""
    import io, contextlib, sys
    stdout_buf = io.StringIO()
    stderr_buf = io.StringIO()
    old_argv = sys.argv
    old_cwd = os.getcwd()
    try:
        os.chdir(cwd)
        sys.argv = [module_name] + list(args)
        mod = __import__(module_name)
        # Find the main/cli function
        entry = None
        for attr in ("main", "cli", "run", "app"):
            if hasattr(mod, attr) and callable(getattr(mod, attr)):
                entry = getattr(mod, attr)
                break
        if entry is None:
            # Try __main__ submodule
            try:
                main_mod = __import__(f"{module_name}.__main__", fromlist=["main"])
                for attr in ("main", "cli", "run"):
                    if hasattr(main_mod, attr) and callable(getattr(main_mod, attr)):
                        entry = getattr(main_mod, attr)
                        break
            except ImportError:
                pass
        if entry is None:
            return f"{module_name}: no CLI entry point found"
        with contextlib.redirect_stdout(stdout_buf), contextlib.redirect_stderr(stderr_buf):
            try:
                entry()
            except SystemExit:
                pass
        out = stdout_buf.getvalue()
        err = stderr_buf.getvalue()
        return (out + ("\n" + err if err else "")).strip()
    except Exception as e:
        return f"{module_name}: {e}"
    finally:
        sys.argv = old_argv
        try:
            os.chdir(old_cwd)
        except Exception:
            pass


def _cli_jq(args, cwd, **kw):
    """Lightweight jq-like JSON query. Supports .key, .key.sub, .[0], .[], keys, length."""
    import json
    stdin = kw.get("stdin", "")
    if not args:
        return "Usage: jq <filter> [file]"
    filt = args[0]
    # Read input from file arg or stdin
    data_str = stdin
    if len(args) > 1:
        fp = os.path.join(cwd, args[1])
        if os.path.isfile(fp):
            with open(fp, "r") as f:
                data_str = f.read()
    if not data_str.strip():
        return "jq: no input"
    try:
        data = json.loads(data_str)
    except json.JSONDecodeError as e:
        return f"jq: parse error: {e}"

    def _apply(obj, expr):
        expr = expr.strip()
        if expr == ".":
            return obj
        if expr == "keys":
            if isinstance(obj, dict):
                return list(obj.keys())
            return f"jq: cannot get keys of {type(obj).__name__}"
        if expr == "length":
            return len(obj)
        if expr == "type":
            return type(obj).__name__
        if expr == "values":
            if isinstance(obj, dict):
                return list(obj.values())
            return obj
        if expr == "flatten":
            if isinstance(obj, list):
                result = []
                for item in obj:
                    if isinstance(item, list):
                        result.extend(item)
                    else:
                        result.append(item)
                return result
            return obj
        if expr == "reverse":
            if isinstance(obj, list):
                return list(reversed(obj))
            return obj
        if expr == "sort":
            if isinstance(obj, list):
                return sorted(obj, key=str)
            return obj
        if expr == "unique":
            if isinstance(obj, list):
                seen = []
                for x in obj:
                    if x not in seen:
                        seen.append(x)
                return seen
            return obj
        # .[] - iterate array/object values
        if expr == ".[]":
            if isinstance(obj, list):
                return obj
            if isinstance(obj, dict):
                return list(obj.values())
            return obj
        # .[N] - array index
        m = _re.match(r'^\.\[(\d+)\](.*)$', expr)
        if m:
            idx = int(m.group(1))
            rest = m.group(2)
            if isinstance(obj, list) and idx < len(obj):
                result = obj[idx]
                return _apply(result, rest) if rest else result
            return None
        # .key or .key.sub.sub
        if expr.startswith("."):
            parts = expr[1:].split(".")
            cur = obj
            for p in parts:
                if not p:
                    continue
                # Handle [N] suffix: .key[0]
                bracket = _re.match(r'^(\w+)\[(\d+)\]$', p)
                if bracket:
                    key, idx = bracket.group(1), int(bracket.group(2))
                    if isinstance(cur, dict) and key in cur:
                        cur = cur[key]
                        if isinstance(cur, list) and idx < len(cur):
                            cur = cur[idx]
                        else:
                            return None
                    else:
                        return None
                elif isinstance(cur, dict) and p in cur:
                    cur = cur[p]
                elif isinstance(cur, list):
                    # Map over array: .name on [{name:"a"},{name:"b"}] -> ["a","b"]
                    cur = [item.get(p) if isinstance(item, dict) else None for item in cur]
                else:
                    return None
            return cur
        # select(expr) - basic filtering
        m = _re.match(r'^select\(\.(\w+)\s*(==|!=|>|<|>=|<=)\s*(.+)\)$', expr)
        if m:
            key, op, val = m.group(1), m.group(2), m.group(3).strip().strip('"\'')
            try:
                val = json.loads(val)
            except (json.JSONDecodeError, ValueError):
                pass
            ops = {"==": lambda a,b: a==b, "!=": lambda a,b: a!=b,
                   ">": lambda a,b: a>b, "<": lambda a,b: a<b,
                   ">=": lambda a,b: a>=b, "<=": lambda a,b: a<=b}
            if isinstance(obj, list):
                return [item for item in obj if isinstance(item, dict) and ops.get(op, lambda a,b: False)(item.get(key), val)]
            if isinstance(obj, dict):
                return obj if ops.get(op, lambda a,b: False)(obj.get(key), val) else None
        return f"jq: unsupported filter: {expr}"

    # Handle pipe chains in jq filter: .data | .[] | .name
    parts = [p.strip() for p in filt.split("|")]
    result = data
    for part in parts:
        if isinstance(result, list) and part.startswith(".") and part != ".[]" and not part.startswith(".["):
            # Auto-map over arrays
            result = [_apply(item, part) for item in result]
        else:
            result = _apply(result, part)

    if isinstance(result, (dict, list)):
        return json.dumps(result, indent=2, ensure_ascii=False)
    if result is None:
        return "null"
    return str(result)


def _cli_sqlite3(args, cwd, **kw):
    """SQLite3 CLI interface. Usage: sqlite3 <db> <sql> or sqlite3 <db> '.tables'"""
    import sqlite3
    stdin = kw.get("stdin", "")
    if not args:
        return "Usage: sqlite3 <database> [sql|.command]"
    db_path = os.path.join(cwd, args[0])
    sql = " ".join(args[1:]) if len(args) > 1 else stdin.strip()
    if not sql:
        return "sqlite3: no SQL provided"
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        # Handle dot-commands
        if sql.startswith("."):
            cmd = sql.split()[0]
            if cmd == ".tables":
                cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
                return "\n".join(row[0] for row in cur.fetchall())
            elif cmd == ".schema":
                table = sql.split()[1] if len(sql.split()) > 1 else None
                if table:
                    cur.execute("SELECT sql FROM sqlite_master WHERE name=?", (table,))
                else:
                    cur.execute("SELECT sql FROM sqlite_master WHERE sql IS NOT NULL ORDER BY name")
                return "\n".join(row[0] for row in cur.fetchall() if row[0])
            elif cmd == ".databases":
                return f"main: {db_path}"
            elif cmd == ".dump":
                lines = []
                for line in conn.iterdump():
                    lines.append(line)
                return "\n".join(lines)
            elif cmd == ".indices" or cmd == ".indexes":
                table = sql.split()[1] if len(sql.split()) > 1 else None
                if table:
                    cur.execute("SELECT name FROM sqlite_master WHERE type='index' AND tbl_name=?", (table,))
                else:
                    cur.execute("SELECT name FROM sqlite_master WHERE type='index' ORDER BY name")
                return "\n".join(row[0] for row in cur.fetchall())
            else:
                return f"sqlite3: unknown command: {cmd}"
        # Execute SQL
        # Support multiple statements separated by ;
        statements = [s.strip() for s in sql.split(";") if s.strip()]
        results = []
        for stmt in statements:
            cur.execute(stmt)
            if cur.description:
                cols = [d[0] for d in cur.description]
                rows = cur.fetchall()
                results.append("|".join(cols))
                for row in rows:
                    results.append("|".join(str(v) for v in row))
            elif cur.rowcount >= 0:
                results.append(f"({cur.rowcount} rows affected)")
        conn.commit()
        conn.close()
        return "\n".join(results)
    except Exception as e:
        return f"sqlite3: {e}"


def _cli_json_pp(args, cwd, **kw):
    """Pretty-print JSON (like json_pp or python -m json.tool)."""
    import json
    stdin = kw.get("stdin", "")
    data_str = stdin
    if args:
        fp = os.path.join(cwd, args[0])
        if os.path.isfile(fp):
            with open(fp, "r") as f:
                data_str = f.read()
    if not data_str.strip():
        return "json_pp: no input"
    try:
        obj = json.loads(data_str)
        indent = 2
        for i, a in enumerate(args):
            if a == "--indent" and i + 1 < len(args):
                indent = int(args[i + 1])
        return json.dumps(obj, indent=indent, ensure_ascii=False)
    except Exception as e:
        return f"json_pp: {e}"


def _cli_bc(args, cwd, **kw):
    """Basic calculator (like bc). Evaluates math expressions."""
    import math
    stdin = kw.get("stdin", "")
    expr = " ".join(args) if args else stdin.strip()
    if not expr:
        return "bc: no expression"
    # Replace common bc-isms
    expr = expr.replace("^", "**").replace("sqrt", "math.sqrt")
    expr = expr.replace("s(", "math.sin(").replace("c(", "math.cos(")
    expr = expr.replace("l(", "math.log(").replace("a(", "math.atan(")
    try:
        result = eval(expr, {"__builtins__": {}, "math": math, "pi": math.pi, "e": math.e})
        return str(result)
    except Exception as e:
        return f"bc: {e}"


def _cli_htop(args, cwd, **kw):
    """Show process/resource info (htop-like). Reports Python runtime stats."""
    import resource, platform, threading, gc
    try:
        usage = resource.getrusage(resource.RUSAGE_SELF)
        mem_mb = usage.ru_maxrss / (1024 * 1024) if sys.platform == "darwin" else usage.ru_maxrss / 1024
        lines = [
            f"Pegasus Agent Process Info",
            f"{'='*40}",
            f"Platform:      {platform.system()} {platform.machine()}",
            f"Python:        {platform.python_version()}",
            f"Threads:       {threading.active_count()}",
            f"Peak Memory:   {mem_mb:.1f} MB",
            f"User CPU:      {usage.ru_utime:.2f}s",
            f"System CPU:    {usage.ru_stime:.2f}s",
            f"GC Objects:    {len(gc.get_objects())}",
            f"GC Collections: gen0={gc.get_count()[0]} gen1={gc.get_count()[1]} gen2={gc.get_count()[2]}",
        ]
        # Show loaded modules count
        lines.append(f"Modules:       {len(sys.modules)}")
        # Show workspace size
        total_files = 0
        total_size = 0
        for root, dirs, files in os.walk(SANDBOX_ROOT):
            total_files += len(files)
            for f in files:
                try:
                    total_size += os.path.getsize(os.path.join(root, f))
                except OSError:
                    pass
            if total_files > 10000:
                break
        lines.append(f"Workspace:     {total_files} files, {total_size / (1024*1024):.1f} MB")
        return "\n".join(lines)
    except Exception as e:
        return f"htop: {e}"


def _cli_tree(args, cwd, **kw):
    """Show directory tree (like the `tree` command)."""
    target = os.path.join(cwd, args[0]) if args else cwd
    if not os.path.isdir(target):
        return f"tree: {target}: not a directory"
    max_depth = 4
    show_hidden = False
    for i, a in enumerate(args):
        if a == "-L" and i + 1 < len(args):
            try:
                max_depth = int(args[i + 1])
            except ValueError:
                pass
        if a == "-a":
            show_hidden = True
    lines = [os.path.basename(target) or target]
    dir_count = 0
    file_count = 0
    def _walk(path, prefix, depth):
        nonlocal dir_count, file_count
        if depth >= max_depth:
            return
        try:
            entries = sorted(os.listdir(path))
        except PermissionError:
            return
        if not show_hidden:
            entries = [e for e in entries if not e.startswith(".")]
        for i, entry in enumerate(entries):
            is_last = (i == len(entries) - 1)
            connector = "└── " if is_last else "├── "
            full = os.path.join(path, entry)
            lines.append(f"{prefix}{connector}{entry}")
            if os.path.isdir(full):
                dir_count += 1
                ext = "    " if is_last else "│   "
                _walk(full, prefix + ext, depth + 1)
            else:
                file_count += 1
    _walk(target, "", 0)
    lines.append(f"\n{dir_count} directories, {file_count} files")
    return "\n".join(lines)


def _cli_nc(args, cwd, **kw):
    """Netcat-like: nc host port sends stdin or tests connectivity."""
    stdin = kw.get("stdin", "")
    if len(args) < 2:
        return "Usage: nc <host> <port>"
    host, port = args[0], int(args[1])
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(5)
        s.connect((host, port))
        if stdin:
            s.sendall(stdin.encode())
        s.shutdown(socket.SHUT_WR)
        resp = b""
        while True:
            chunk = s.recv(4096)
            if not chunk:
                break
            resp += chunk
        s.close()
        return resp.decode("utf-8", errors="replace")
    except Exception as e:
        return f"nc: {e}"


def _cli_discover_entry_point(cmd, args, cwd):
    """Auto-discover console_scripts entry points from pip-installed packages.
    Checks dist-info directories for entry_points.txt or metadata."""
    # Map common CLI names to their Python module
    _KNOWN_MODULES = {
        "yt-dlp": "yt_dlp",
        "youtube-dl": "youtube_dl",
        "black": "black",
        "ruff": "ruff",
        "isort": "isort",
        "mypy": "mypy",
        "flake8": "flake8",
        "pytest": "pytest",
        "httpie": "httpie",
        "http": "httpie",
        "requests": "requests",
        "scrapy": "scrapy",
        "flask": "flask",
        "gunicorn": "gunicorn",
        "uvicorn": "uvicorn",
        "streamlit": "streamlit",
        "jupyter": "jupyter",
        "ipython": "IPython",
        "rich": "rich",
        "typer": "typer",
        "cookiecutter": "cookiecutter",
        "ansible": "ansible",
        "fabric": "fabric",
    }

    module_name = _KNOWN_MODULES.get(cmd, cmd.replace("-", "_"))

    # Try importing and running the module's CLI
    try:
        mod = __import__(module_name)
        return _cli_run_module(module_name, args, cwd)
    except ImportError:
        pass

    # Scan dist-info for entry points
    for search_dir in [_PACKAGES_DIR, _CUSTOM_PACKAGES_DIR]:
        if not os.path.isdir(search_dir):
            continue
        for item in os.listdir(search_dir):
            if item.endswith(".dist-info"):
                ep_file = os.path.join(search_dir, item, "entry_points.txt")
                if os.path.isfile(ep_file):
                    with open(ep_file, "r") as f:
                        content = f.read()
                    # Parse entry_points.txt for console_scripts
                    in_console = False
                    for line in content.splitlines():
                        line = line.strip()
                        if line == "[console_scripts]":
                            in_console = True
                            continue
                        if line.startswith("["):
                            in_console = False
                            continue
                        if in_console and "=" in line:
                            name, target = line.split("=", 1)
                            if name.strip() == cmd:
                                # target is like "module:func"
                                mod_path, func_name = target.strip().split(":")
                                try:
                                    mod = __import__(mod_path, fromlist=[func_name])
                                    entry = getattr(mod, func_name)
                                    import io, contextlib
                                    old_argv = sys.argv
                                    sys.argv = [cmd] + list(args)
                                    stdout_buf = io.StringIO()
                                    stderr_buf = io.StringIO()
                                    try:
                                        with contextlib.redirect_stdout(stdout_buf), contextlib.redirect_stderr(stderr_buf):
                                            try:
                                                entry()
                                            except SystemExit:
                                                pass
                                        return (stdout_buf.getvalue() + stderr_buf.getvalue()).strip()
                                    finally:
                                        sys.argv = old_argv
                                except Exception as e:
                                    return f"{cmd}: {e}"
    return None  # Not found


# Built-in CLI adapters — tools with Python-native implementations
_CLI_ADAPTERS = {
    "jq": _cli_jq,
    "json_pp": _cli_json_pp,
    "json_tool": _cli_json_pp,
    "sqlite3": _cli_sqlite3,
    "sqlite": _cli_sqlite3,
    "bc": _cli_bc,
    "calc": _cli_bc,
    "htop": _cli_htop,
    "top": _cli_htop,
    "tree": _cli_tree,
    "nc": _cli_nc,
    "netcat": _cli_nc,
    "ncat": _cli_nc,
}


# Command dispatch table — 70+ commands
_SHELL_COMMANDS = {
    "ls": _shell_ls,
    "cat": _shell_cat_n,
    "head": _shell_head,
    "tail": _shell_tail,
    "grep": _shell_grep,
    "egrep": _shell_grep,
    "fgrep": lambda a, c, **kw: _shell_grep(["-F"] + a, c, **kw),
    "find": _shell_find,
    "wc": _shell_wc,
    "mkdir": _shell_mkdir,
    "rm": _shell_rm,
    "rmdir": lambda args, cwd, **kw: (os.rmdir(os.path.join(cwd, args[0])) or "") if args else "rmdir: missing dir",
    "cp": _shell_cp,
    "mv": _shell_mv,
    "echo": _shell_echo,
    "printf": _shell_echo,
    "pwd": _shell_pwd,
    "curl": _shell_curl,
    "wget": _shell_wget,
    "touch": _shell_touch,
    "du": _shell_du,
    "which": _shell_which,
    "env": _shell_env,
    "printenv": _shell_env,
    "date": _shell_date,
    "sort": _shell_sort,
    "uniq": _shell_uniq,
    "sed": _shell_sed,
    "awk": _shell_awk,
    "cut": _shell_cut,
    "tr": _shell_tr,
    "tee": _shell_tee,
    "xargs": _shell_xargs,
    "diff": _shell_diff,
    "tar": _shell_tar,
    "zip": _shell_zip,
    "unzip": _shell_unzip,
    "gzip": _shell_gzip,
    "gunzip": lambda a, c, **kw: _shell_gzip(["-d"] + a, c),
    "basename": _shell_basename,
    "dirname": _shell_dirname,
    "realpath": _shell_realpath,
    "readlink": _shell_readlink,
    "ln": _shell_ln,
    "stat": _shell_stat,
    "file": _shell_file,
    "md5sum": _shell_md5sum,
    "md5": _shell_md5sum,
    "sha256sum": _shell_sha256sum,
    "shasum": _shell_sha256sum,
    "base64": _shell_base64_cmd,
    "xxd": _shell_xxd,
    "hexdump": _shell_xxd,
    "seq": _shell_seq,
    "sleep": _shell_sleep,
    "tac": _shell_tac,
    "rev": _shell_rev,
    "nl": _shell_nl,
    "paste": _shell_paste,
    "column": _shell_column,
    "cal": _shell_cal,
    "expr": _shell_expr,
    "test": _shell_test,
    "[": _shell_test,
    "export": _shell_export,
    "set": lambda a, c, **kw: "\n".join(f"{k}={v}" for k, v in sorted(os.environ.items())),
    "unset": lambda a, c, **kw: [os.environ.pop(x, None) for x in a] and "" or "",
    "chmod": _shell_chmod,
    "chown": _shell_chown,
    "whoami": _shell_whoami,
    "hostname": _shell_hostname,
    "uname": _shell_uname,
    "id": lambda a, c, **kw: "uid=501(mobile) gid=501(mobile)",
    "uptime": lambda a, c, **kw: f"up {int(((__import__('time').time() - __import__('time').time() % 86400) / 3600))} hours",
    "true": lambda a, c, **kw: "",
    "false": lambda a, c, **kw: "",
    "clear": lambda a, c, **kw: "",
    "reset": lambda a, c, **kw: "",
    "exit": lambda a, c, **kw: "",
    "logout": lambda a, c, **kw: "",
    "history": lambda a, c, **kw: "(no history in emulated shell)",
    "alias": lambda a, c, **kw: "(aliases not supported)",
    "type": lambda a, c, **kw: "\n".join(f"{x} is a shell emulated command" if x in _SHELL_COMMANDS else f"{x}: not found" for x in a),
    "nohup": lambda a, c, **kw: _exec_single_command(" ".join(a), c),
    "time": lambda a, c, **kw: _exec_single_command(" ".join(a), c),
    "nice": lambda a, c, **kw: _exec_single_command(" ".join(a), c),
}


def _parse_shell_line(command_str):
    """Minimal shell argument parser (handles quotes and pipes)."""
    import shlex
    try:
        return shlex.split(command_str)
    except ValueError:
        return command_str.split()


def shell_exec(command: str, timeout: int = 30) -> dict:
    """Execute a shell command emulated via Python.

    Supports 70+ Unix commands implemented in pure Python.
    Pipes (|), chains (&&, ;), redirects (>, >>), and subshells $() work.
    Commands like 'python' and 'pip install' are routed to python_exec and pip_install.
    """
    command = command.strip()
    if not command:
        return {"stdout": "", "stderr": "", "returncode": 0}

    cwd = SANDBOX_ROOT

    try:
        # Handle && chains FIRST (highest level separator)
        if "&&" in command:
            parts = [s.strip() for s in command.split("&&")]
            outputs = []
            for part in parts:
                r = _run_single_statement(part, cwd)
                if r:
                    outputs.append(r)
            return {"stdout": "\n".join(outputs)[:50_000], "stderr": "", "returncode": 0}

        # Handle ; chains (but not inside quotes)
        if ";" in command and "'" not in command and '"' not in command:
            parts = [s.strip() for s in command.split(";") if s.strip()]
            outputs = []
            for part in parts:
                r = _run_single_statement(part, cwd)
                if r:
                    outputs.append(r)
            return {"stdout": "\n".join(outputs)[:50_000], "stderr": "", "returncode": 0}

        result = _run_single_statement(command, cwd)
        return {"stdout": result[:50_000], "stderr": "", "returncode": 0}

    except Exception as e:
        return {"stdout": "", "stderr": str(e)[:10_000], "returncode": 1}


def _run_single_statement(command, cwd):
    """Handle a single statement: redirects first, then pipes."""
    # Handle >> BEFORE > (>> contains >)
    if ">>" in command:
        parts = command.split(">>", 1)
        cmd_part = parts[0].strip()
        file_part = parts[1].strip()
        result = _run_with_pipes(cmd_part, cwd)
        fp = os.path.join(cwd, file_part)
        os.makedirs(os.path.dirname(fp) or cwd, exist_ok=True)
        with open(fp, "a", encoding="utf-8") as f:
            f.write(result)
        return ""
    if ">" in command:
        parts = command.split(">", 1)
        cmd_part = parts[0].strip()
        file_part = parts[1].strip()
        result = _run_with_pipes(cmd_part, cwd)
        fp = os.path.join(cwd, file_part)
        os.makedirs(os.path.dirname(fp) or cwd, exist_ok=True)
        with open(fp, "w", encoding="utf-8") as f:
            f.write(result)
        return ""
    return _run_with_pipes(command, cwd)


def _run_with_pipes(command, cwd):
    """Handle pipes by chaining commands, passing stdout as stdin."""
    if "|" in command:
        segments = [s.strip() for s in command.split("|")]
        pipe_data = ""
        for seg in segments:
            pipe_data = _exec_single_command(seg, cwd, stdin=pipe_data)
        return pipe_data
    return _exec_single_command(command, cwd)


def _exec_single_command(command_str, cwd, stdin=""):
    """Execute a single command (no pipes/chains). Passes stdin to handlers that accept it."""
    tokens = _parse_shell_line(command_str)
    if not tokens:
        return stdin  # pass-through for empty segments

    cmd = tokens[0]
    args = tokens[1:]

    # Handle cd
    if cmd == "cd":
        return ""

    # Handle variable assignment: VAR=value
    if "=" in cmd and not cmd.startswith("-") and _re.match(r'^[A-Za-z_]\w*=', cmd):
        k, v = cmd.split("=", 1)
        os.environ[k] = v
        return ""

    # Route python/python3 to python_exec
    if cmd in ("python", "python3"):
        if args and args[0] == "-c" and len(args) > 1:
            code = " ".join(args[1:])
            result = python_exec(code)
            return result.get("stdout", "") + result.get("error", "")
        elif args and not args[0].startswith("-"):
            fp = os.path.join(cwd, args[0])
            if os.path.isfile(fp):
                with open(fp, "r") as f:
                    code = f.read()
                result = python_exec(code)
                return result.get("stdout", "") + result.get("error", "")
            return f"python: can't open file '{args[0]}'"
        return "Python 3 (embedded, iOS)"

    # Route pip install to pip_install tool
    if cmd in ("pip", "pip3"):
        if args and args[0] == "install":
            packages = [a for a in args[1:] if not a.startswith("-")]
            results = []
            for pkg in packages:
                r = pip_install(pkg)
                results.append(f"{pkg}: {r.get('status', r.get('error', 'unknown'))}")
            return "\n".join(results)
        if args and args[0] == "list":
            installed = []
            if os.path.isdir(_PACKAGES_DIR):
                for item in sorted(os.listdir(_PACKAGES_DIR)):
                    if item.endswith(".dist-info"):
                        installed.append(item.replace(".dist-info", ""))
            return "\n".join(installed) if installed else "(no packages installed)"
        if args and args[0] == "show":
            pkg = args[1] if len(args) > 1 else ""
            return f"Package: {pkg}\nLocation: {_PACKAGES_DIR}"
        if args and args[0] == "freeze":
            installed = []
            if os.path.isdir(_PACKAGES_DIR):
                for item in sorted(os.listdir(_PACKAGES_DIR)):
                    if item.endswith(".dist-info"):
                        installed.append(item.replace(".dist-info", "").replace("-", "==", 1))
            return "\n".join(installed)
        return "pip (emulated) - commands: install, list, show, freeze"

    # Look up in command table — pass stdin via kwargs for handlers that accept it
    handler = _SHELL_COMMANDS.get(cmd)
    if handler:
        import inspect
        try:
            sig = inspect.signature(handler)
            if "stdin" in sig.parameters or any(
                p.kind == inspect.Parameter.VAR_KEYWORD for p in sig.parameters.values()
            ):
                return handler(args, cwd, stdin=stdin) or ""
            else:
                return handler(args, cwd) or ""
        except (TypeError, ValueError):
            return handler(args, cwd) or ""

    # Check CLI adapters (jq, sqlite3, tree, bc, htop, etc.)
    cli_handler = _CLI_ADAPTERS.get(cmd)
    if cli_handler:
        import inspect
        try:
            sig = inspect.signature(cli_handler)
            if "stdin" in sig.parameters or any(
                p.kind == inspect.Parameter.VAR_KEYWORD for p in sig.parameters.values()
            ):
                return cli_handler(args, cwd, stdin=stdin) or ""
            else:
                return cli_handler(args, cwd) or ""
        except (TypeError, ValueError):
            return cli_handler(args, cwd) or ""

    # Try auto-discovering CLI entry points from pip-installed packages
    discovered = _cli_discover_entry_point(cmd, args, cwd)
    if discovered is not None:
        return discovered

    # Fallback: try running as Python code
    result = python_exec(command_str)
    stdout = result.get("stdout", "")
    error = result.get("error", "")
    if error and not stdout:
        return f"shell: {cmd}: command not found"
    return stdout


registry.register(
    name="shell_exec",
    description="Execute shell commands. Supports 80+ Unix commands (ls, cat, grep, find, head, tail, cp, mv, rm, mkdir, curl, wget, sed, etc.), CLI tools (jq, sqlite3, tree, htop, bc, nc), python, pip install, and auto-discovers pip-installed package CLIs (yt-dlp, black, ruff, etc.). Pipes (|), chains (&&, ;), and redirects (>, >>) work.",
    parameters={
        "type": "object",
        "properties": {
            "command": {"type": "string", "description": "Shell command to run (e.g. 'ls -la', 'grep -r pattern .', 'curl https://...')"},
            "timeout": {"type": "integer", "description": "Timeout in seconds (default 30)"},
        },
        "required": ["command"],
    },
    handler=shell_exec,
    category="system",
)


# ---- Python Execution ----

# Persistent namespace so imports survive between calls.
# Importing openpyxl/pandas/etc once instead of every call saves seconds.
_EXEC_GLOBALS = {"__builtins__": __builtins__}

# Pre-load common stdlib modules into the persistent namespace
for _mod_name in ("json", "math", "os", "re", "statistics", "collections",
                  "itertools", "functools", "datetime", "pathlib", "csv"):
    try:
        import importlib as _il
        _EXEC_GLOBALS[_mod_name] = _il.import_module(_mod_name)
    except ImportError:
        pass


def python_exec(code: str) -> dict:
    """Execute a Python code snippet and return its output.
    Working directory is the workspace root. All workspace files are accessible.
    Imports persist between calls for speed.
    """
    import io
    import contextlib
    import threading

    # Compile first — catches syntax errors fast, runs faster than raw string exec
    try:
        compiled = compile(code, "<agent>", "exec")
    except SyntaxError as e:
        return {"error": f"SyntaxError: {e}"}

    stdout_capture = io.StringIO()
    stderr_capture = io.StringIO()
    exec_error = [None]
    old_cwd = os.getcwd()
    try:
        os.chdir(SANDBOX_ROOT)
    except Exception:
        pass

    def _run():
        try:
            with contextlib.redirect_stdout(stdout_capture), contextlib.redirect_stderr(stderr_capture):
                exec(compiled, _EXEC_GLOBALS)
        except Exception:
            exec_error[0] = traceback.format_exc()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    out_limit = 100000 if _CLOUD_MODE else 5000
    exec_timeout = 3600 if _CLOUD_MODE else 300  # 1 hour cloud, 5 min local
    # Poll in 2s intervals so we can check for interrupts instead of blocking
    interrupt_file = os.path.join(os.environ.get("TMPDIR", "/tmp"), "pegasus_interrupt")
    elapsed = 0
    while elapsed < exec_timeout and thread.is_alive():
        thread.join(timeout=2)
        elapsed += 2
        # Check for interrupt — allows user to cancel stuck execution
        # Don't remove the file — let the agent loop see it too and abort
        if os.path.exists(interrupt_file):
            break

    try:
        os.chdir(old_cwd)
    except Exception:
        pass

    if thread.is_alive():
        was_interrupted = elapsed < exec_timeout  # broke out early = interrupted
        # Reset execution namespace to prevent corrupted state from hung thread
        global _EXEC_GLOBALS
        _EXEC_GLOBALS = {"__builtins__": __builtins__}
        for _rmod in ("json", "math", "os", "re", "statistics", "collections",
                      "itertools", "functools", "datetime", "pathlib", "csv"):
            try:
                import importlib as _il2
                _EXEC_GLOBALS[_rmod] = _il2.import_module(_rmod)
            except ImportError:
                pass
        msg = "Execution interrupted by user." if was_interrupted else f"Execution timed out after {elapsed} seconds."
        return {
            "stdout": stdout_capture.getvalue()[:out_limit],
            "error": msg,
        }

    if exec_error[0]:
        return {
            "stdout": stdout_capture.getvalue()[:out_limit],
            "error": exec_error[0][:out_limit],
        }

    return {
        "stdout": stdout_capture.getvalue()[:out_limit],
        "stderr": stderr_capture.getvalue()[:20000],
        "result": str(_EXEC_GLOBALS.get("result", ""))[:out_limit],
    }


registry.register(
    name="python_exec",
    description="Execute Python code in the workspace directory. All workspace files are accessible. Set 'result' variable to return a value. Use pip_install first if you need a package.",
    parameters={
        "type": "object",
        "properties": {
            "code": {"type": "string", "description": "Python code to execute"},
        },
        "required": ["code"],
    },
    handler=python_exec,
    category="system",
)


# ---- Excel Reader ----

def excel_read(file: str, sheet: str = "", range: str = "", max_rows: int = 200) -> dict:
    """Read an Excel file efficiently using read_only streaming mode.
    Handles large files (30MB+) that would timeout in python_exec.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed. Call pip_install('openpyxl') first."}

    file_path = os.path.join(SANDBOX_ROOT, file) if not os.path.isabs(file) else file
    if not os.path.isfile(file_path):
        return {"error": f"File not found: {file}"}

    try:
        # read_only=True streams rows — 10x faster, uses fraction of memory
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        if sheet:
            if sheet not in sheet_names:
                wb.close()
                return {"error": f"Sheet '{sheet}' not found. Available: {sheet_names}"}
            ws = wb[sheet]
        else:
            ws = wb.active

        # Parse range if given (e.g. "A1:F50")
        min_row, max_row_limit, min_col, max_col = 1, max_rows, None, None
        if range:
            m = _re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range.upper())
            if m:
                from openpyxl.utils import column_index_from_string
                min_col = column_index_from_string(m.group(1))
                min_row = int(m.group(2))
                max_col = column_index_from_string(m.group(3))
                max_row_limit = int(m.group(4)) - min_row + 1

        rows = []
        count = 0
        for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col, values_only=True):
            # Convert to strings, replace None with empty
            rows.append([str(c) if c is not None else "" for c in row])
            count += 1
            if count >= max_row_limit:
                break

        wb.close()

        # Format as readable table
        if not rows:
            return {"sheet": ws.title, "sheets": sheet_names, "data": "(empty)"}

        # Find max width per column for alignment
        col_widths = [0] * len(rows[0])
        for row in rows[:50]:  # Sample first 50 for widths
            for i, cell in enumerate(row):
                if i < len(col_widths):
                    col_widths[i] = max(col_widths[i], min(len(cell), 20))

        lines = []
        for row in rows:
            parts = []
            for i, cell in enumerate(row):
                val = cell[:20]  # Truncate wide cells
                parts.append(val)
            lines.append(" | ".join(parts))

        table = "\n".join(lines)
        # Cap total output
        cap = 50000 if _CLOUD_MODE else 5000
        if len(table) > cap:
            table = table[:cap] + f"\n[truncated, showing {count} rows]"

        return {
            "sheet": ws.title,
            "sheets": sheet_names,
            "rows": count,
            "data": table,
        }
    except Exception:
        return {"error": traceback.format_exc()[:3000]}


registry.register(
    name="excel_read",
    description="Read Excel files efficiently. Handles large files (30MB+). Specify sheet name and cell range to read only what you need. Returns formatted table.",
    parameters={
        "type": "object",
        "properties": {
            "file": {"type": "string", "description": "Excel file path (relative to workspace or absolute)"},
            "sheet": {"type": "string", "description": "Sheet name (default: active sheet)"},
            "range": {"type": "string", "description": "Cell range like 'A1:F50' (default: all rows up to max_rows)"},
            "max_rows": {"type": "integer", "description": "Max rows to read (default: 200)"},
        },
        "required": ["file"],
    },
    handler=excel_read,
    category="system",
)


# ---- Web Fetch ----

def web_fetch(url: str) -> dict:
    """Fetch content from a URL and return clean text (HTML stripped)."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Pegasus/1.0"})
        with urllib.request.urlopen(req, timeout=15, context=_ssl_ctx) as resp:
            read_limit = 500_000 if _CLOUD_MODE else 200_000
            raw = resp.read(read_limit).decode("utf-8", errors="replace")
        # Strip HTML to give the model clean text
        content = _strip_html(raw)
        # Only truncate for on-device models with limited context
        if not _CLOUD_MODE:
            max_len = 3000
            if len(content) > max_len:
                content = content[:max_len] + "\n[Truncated]"
        return {"url": url, "status": resp.status, "content": content}
    except Exception as e:
        return {"error": str(e)}


registry.register(
    name="web_fetch",
    description="Fetch content from a URL and return the text. Use this to scrape web pages.",
    parameters={
        "type": "object",
        "properties": {
            "url": {"type": "string", "description": "URL to fetch"},
        },
        "required": ["url"],
    },
    handler=web_fetch,
    category="web",
)


def web_search(query: str, num_results: int = 5) -> dict:
    """Search the web using DuckDuckGo and return results."""
    import html
    import re
    try:
        from urllib.parse import quote, unquote
        encoded_query = quote(query)
        url = f"https://html.duckduckgo.com/html/?q={encoded_query}"
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15"
        })
        with urllib.request.urlopen(req, timeout=15, context=_ssl_ctx) as resp:
            page = resp.read().decode("utf-8", errors="replace")

        results = []
        # Parse DuckDuckGo HTML results
        result_blocks = re.findall(
            r'<a rel="nofollow" class="result__a" href="([^"]*)"[^>]*>(.*?)</a>.*?'
            r'<a class="result__snippet"[^>]*>(.*?)</a>',
            page, re.DOTALL
        )
        for href, title, snippet in result_blocks[:num_results]:
            # Clean up DuckDuckGo redirect URLs
            if "uddg=" in href:
                match = re.search(r'uddg=([^&]+)', href)
                if match:
                    href = unquote(match.group(1))
            clean_title = html.unescape(re.sub(r'<[^>]+>', '', title)).strip()
            clean_snippet = html.unescape(re.sub(r'<[^>]+>', '', snippet)).strip()
            # Remove any image URLs from snippets
            clean_snippet = _re.sub(r'https?://\S+\.(png|jpg|jpeg|gif|svg|webp)\S*', '', clean_snippet, flags=_re.IGNORECASE)
            results.append({
                "title": clean_title,
                "url": href,
                "snippet": clean_snippet,
            })

        # Limit total result size for on-device models only
        if not _CLOUD_MODE:
            total = json.dumps(results, ensure_ascii=False)
            if len(total) > 4000:
                results = results[:3]
        return {"query": query, "results": results}
    except Exception as e:
        return {"error": str(e)}


registry.register(
    name="web_search",
    description="Search the web using DuckDuckGo. Returns titles, URLs, and snippets for top results.",
    parameters={
        "type": "object",
        "properties": {
            "query": {"type": "string", "description": "Search query"},
            "num_results": {"type": "integer", "description": "Number of results (default 5)"},
        },
        "required": ["query"],
    },
    handler=web_search,
    category="web",
)


# ---- iOS Native APIs ----
# These work via file-based IPC: Python writes an action request,
# Swift picks it up and executes the native API call, writes result back.

import threading as _threading
_ios_action_lock = _threading.Lock()
_ios_action_counter = 0

def _ios_action(action_type: str, payload: dict, timeout: int = 10) -> dict:
    """Send an action to Swift and wait for result via file IPC.
    Uses unique request IDs so parallel calls don't collide."""
    import time as _t
    global _ios_action_counter

    tmpdir = os.environ.get("TMPDIR", "/tmp")

    # Generate unique request ID
    with _ios_action_lock:
        _ios_action_counter += 1
        req_id = f"{int(_t.time() * 1000)}_{_ios_action_counter}"

    request_file = os.path.join(tmpdir, f"pegasus_ios_action_{req_id}.json")
    response_file = os.path.join(tmpdir, f"pegasus_ios_action_result_{req_id}.json")

    # Write request
    request = {"action": action_type, "payload": payload, "id": req_id}
    with open(request_file, "w", encoding="utf-8") as f:
        json.dump(request, f, ensure_ascii=True)
        f.flush()
        os.fsync(f.fileno())

    # Poll for response
    start = _t.time()
    while _t.time() - start < timeout:
        try:
            with open(response_file, "r", encoding="utf-8") as f:
                raw = f.read()
            if raw and len(raw) > 1:
                result = json.loads(raw)
                try:
                    os.remove(response_file)
                except OSError:
                    pass
                return result
        except (FileNotFoundError, json.JSONDecodeError):
            pass
        _t.sleep(0.1)

    # Clean up request file on timeout
    try:
        os.remove(request_file)
    except OSError:
        pass
    return {"error": f"iOS action '{action_type}' timed out after {timeout}s"}


def ios_action(action: str, **kwargs) -> dict:
    """Unified iOS native API access."""
    return _ios_action(action, kwargs)


registry.register(
    name="ios_action",
    description="""Access iOS native APIs. Actions:
- send_message: {to, body, attachments=[]} - Send iMessage/SMS. To attach files, FIRST find the file (use shell_exec find/ls), then pass full paths in attachments=['/path/to/file.pdf']. Attempts auto-send; falls back to in-app composer
- make_call: {number} - Initiate phone call
- open_url: {url} - Open URL/deep-link (Safari, Maps, Shortcuts, tel:, mailto:)
- notify: {title, body, delay=0} - Local notification
- clipboard: {action='get'|'set', text} - Read/write clipboard
- read_contacts: {search=''} - Search contacts by name
- read_calendar: {days=7} - Get upcoming calendar events
- read_reminders: {list=''} - Get reminders
- set_alarm: {hour, minute, label=''} - Set alarm via Clock app
- get_location: {} - Get current GPS coordinates
- get_battery: {} - Battery level and charging state
- get_device_info: {} - Device model, OS version, storage
- haptic: {style='light'|'medium'|'heavy'|'success'|'warning'|'error'}
- share: {text, url=''} - Open share sheet""",
    parameters={
        "type": "object",
        "properties": {
            "action": {"type": "string", "description": "Action name (send_message, read_contacts, notify, etc.)"},
            "to": {"type": "string", "description": "Recipient (for send_message)"},
            "body": {"type": "string", "description": "Message/notification body"},
            "number": {"type": "string", "description": "Phone number (for make_call)"},
            "service": {"type": "string", "description": "imessage or sms (auto-detected, usually not needed)"},
            "url": {"type": "string", "description": "URL to open or share"},
            "text": {"type": "string", "description": "Text for clipboard/share"},
            "title": {"type": "string", "description": "Notification title"},
            "delay": {"type": "integer", "description": "Notification delay in seconds"},
            "search": {"type": "string", "description": "Contact search query"},
            "days": {"type": "integer", "description": "Days ahead for calendar (default 7)"},
            "list": {"type": "string", "description": "Reminder list name"},
            "hour": {"type": "integer", "description": "Alarm hour (0-23)"},
            "minute": {"type": "integer", "description": "Alarm minute (0-59)"},
            "label": {"type": "string", "description": "Alarm label"},
            "style": {"type": "string", "description": "Haptic style"},
            "attachments": {"type": "array", "items": {"type": "string"}, "description": "File paths to attach (for send_message)"},
        },
        "required": ["action"],
    },
    handler=ios_action,
    category="ios",
)


# ---- OCR (Vision) ----

def ocr_image(path: str) -> dict:
    """Read text from an image using on-device OCR (Apple Vision).

    Supports photos, screenshots, scanned documents, receipts, whiteboards.
    Returns extracted text and individual lines.
    """
    full = os.path.realpath(path)
    if not os.path.isfile(full):
        # Try workspace-relative path
        full = os.path.realpath(os.path.join(SANDBOX_ROOT, path))
    if not os.path.isfile(full):
        return {"error": f"Image not found: {path}"}
    return _ios_action("ocr_image", {"path": full}, timeout=30)


registry.register(
    name="ocr_image",
    description="Extract text from an image using on-device OCR. Works with photos, screenshots, scanned documents, receipts, whiteboards. Returns the recognized text.",
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to image file (jpg, png, heic)"},
        },
        "required": ["path"],
    },
    handler=ocr_image,
    category="vision",
)


# ---- Text-to-Speech ----

def speak(text: str, rate: float = 0.5, language: str = "en-US") -> dict:
    """Speak text aloud using on-device text-to-speech.

    Rate: 0.0 (slowest) to 1.0 (fastest), default 0.5.
    """
    if not text:
        return {"error": "No text to speak"}
    return _ios_action("speak", {"text": text, "rate": rate, "language": language})


def stop_speaking() -> dict:
    """Stop any ongoing text-to-speech."""
    return _ios_action("stop_speaking", {})


registry.register(
    name="speak",
    description="Speak text aloud using on-device text-to-speech. Supports multiple languages. Rate: 0.0 (slowest) to 1.0 (fastest).",
    parameters={
        "type": "object",
        "properties": {
            "text": {"type": "string", "description": "Text to speak"},
            "rate": {"type": "number", "description": "Speech rate 0.0-1.0 (default 0.5)"},
            "language": {"type": "string", "description": "Language code (default 'en-US')"},
        },
        "required": ["text"],
    },
    handler=speak,
    category="voice",
)

registry.register(
    name="stop_speaking",
    description="Stop any ongoing text-to-speech output.",
    parameters={"type": "object", "properties": {}, "required": []},
    handler=stop_speaking,
    category="voice",
)


# ---- Voice Recording & Transcription ----

def voice_record(action: str = "start") -> dict:
    """Start or stop voice recording from the microphone.

    action='start' begins recording, 'stop' ends and returns the audio file path.
    After stopping, use transcribe() to convert speech to text.
    """
    if action == "start":
        return _ios_action("start_recording", {})
    elif action == "stop":
        return _ios_action("stop_recording", {})
    return {"error": f"Unknown action: {action}. Use 'start' or 'stop'."}


def transcribe(path: str = "", language: str = "auto") -> dict:
    """Transcribe an audio file to text using on-device whisper.cpp.

    If no path given, transcribes the last recording from voice_record().
    Supports: WAV (16kHz mono preferred), M4A, CAF.
    """
    return _ios_action("transcribe", {"path": path, "language": language}, timeout=60)


registry.register(
    name="voice_record",
    description="Record audio from the microphone. action='start' to begin, 'stop' to end. After stopping, use transcribe() on the returned path.",
    parameters={
        "type": "object",
        "properties": {
            "action": {"type": "string", "enum": ["start", "stop"], "description": "'start' or 'stop'"},
        },
        "required": ["action"],
    },
    handler=voice_record,
    category="voice",
)

registry.register(
    name="transcribe",
    description="Transcribe speech from an audio file to text using on-device whisper.cpp. Works offline, supports 99 languages.",
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to audio file (WAV/M4A). If empty, uses last recording."},
            "language": {"type": "string", "description": "Language code or 'auto' (default)"},
        },
        "required": [],
    },
    handler=transcribe,
    category="voice",
)


# ---- RAG / Semantic Search ----

import sqlite3 as _sqlite3
import math as _math

_RAG_DB_PATH = os.path.join(
    os.environ.get("PEGASUS_DATA_DIR", os.path.expanduser("~/Documents/pegasus_data")),
    "vector_store.db"
)


def _init_vector_db():
    """Initialize the SQLite vector store if it doesn't exist."""
    conn = _sqlite3.connect(_RAG_DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            chunk_idx INTEGER DEFAULT 0,
            content TEXT NOT NULL,
            embedding TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_source ON documents(source)")
    conn.commit()
    conn.close()


def _cosine_similarity(a, b):
    """Compute cosine similarity between two vectors."""
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = _math.sqrt(sum(x * x for x in a))
    norm_b = _math.sqrt(sum(x * x for x in b))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)


def _get_embedding(text):
    """Get embedding vector for text via iOS NLEmbedding."""
    result = _ios_action("embed_text", {"text": text}, timeout=10)
    if "error" in result:
        return None
    return result.get("vector")


def _chunk_text(text, chunk_size=500, overlap=50):
    """Split text into overlapping chunks."""
    words = text.split()
    chunks = []
    i = 0
    while i < len(words):
        chunk = " ".join(words[i:i + chunk_size])
        chunks.append(chunk)
        i += chunk_size - overlap
    return chunks if chunks else [text]


def rag_index(source: str, content: str = "", chunk_size: int = 500) -> dict:
    """Index a document for semantic search.

    Splits the content into chunks, generates embeddings, and stores them.
    If content is empty, reads from the source path.
    """
    _init_vector_db()

    if not content:
        # Try to read from file
        full = os.path.realpath(os.path.join(SANDBOX_ROOT, source))
        if os.path.isfile(full):
            with open(full, "r", errors="replace") as f:
                content = f.read(500_000)  # 500KB limit
        else:
            return {"error": f"No content provided and file not found: {source}"}

    chunks = _chunk_text(content, chunk_size)
    conn = _sqlite3.connect(_RAG_DB_PATH)

    # Remove old entries for this source
    conn.execute("DELETE FROM documents WHERE source = ?", (source,))

    indexed = 0
    for i, chunk in enumerate(chunks):
        embedding = _get_embedding(chunk[:200])  # Embed first 200 chars of chunk
        emb_json = json.dumps(embedding) if embedding else None
        conn.execute(
            "INSERT INTO documents (source, chunk_idx, content, embedding) VALUES (?, ?, ?, ?)",
            (source, i, chunk, emb_json)
        )
        indexed += 1

    conn.commit()
    conn.close()

    return {"status": "indexed", "source": source, "chunks": indexed}


def rag_search(query: str, top_k: int = 5) -> dict:
    """Search indexed documents by semantic similarity.

    Returns the most relevant chunks from all indexed documents.
    """
    _init_vector_db()

    query_emb = _get_embedding(query)
    if not query_emb:
        return {"error": "Could not generate query embedding"}

    conn = _sqlite3.connect(_RAG_DB_PATH)
    rows = conn.execute(
        "SELECT id, source, chunk_idx, content, embedding FROM documents WHERE embedding IS NOT NULL"
    ).fetchall()
    conn.close()

    if not rows:
        return {"error": "No documents indexed. Use rag_index first."}

    results = []
    for row_id, source, chunk_idx, content, emb_json in rows:
        try:
            emb = json.loads(emb_json)
            sim = _cosine_similarity(query_emb, emb)
            results.append({
                "source": source,
                "chunk": chunk_idx,
                "similarity": round(sim, 4),
                "content": content[:500],
            })
        except (json.JSONDecodeError, TypeError):
            continue

    results.sort(key=lambda x: x["similarity"], reverse=True)
    top = results[:top_k]

    return {
        "status": "ok",
        "query": query,
        "results": top,
        "total_chunks": len(rows),
    }


def rag_list() -> dict:
    """List all indexed documents in the vector store."""
    _init_vector_db()
    conn = _sqlite3.connect(_RAG_DB_PATH)
    rows = conn.execute(
        "SELECT source, COUNT(*) as chunks, MIN(created_at) as indexed_at FROM documents GROUP BY source"
    ).fetchall()
    conn.close()

    sources = [{"source": r[0], "chunks": r[1], "indexed_at": r[2]} for r in rows]
    return {"status": "ok", "documents": sources, "total": len(sources)}


def rag_delete(source: str) -> dict:
    """Remove a document from the vector store."""
    _init_vector_db()
    conn = _sqlite3.connect(_RAG_DB_PATH)
    conn.execute("DELETE FROM documents WHERE source = ?", (source,))
    conn.commit()
    conn.close()
    return {"status": "deleted", "source": source}


registry.register(
    name="rag_index",
    description="Index a document for semantic search (RAG). Splits into chunks, generates embeddings, and stores for later retrieval. Provide either content directly or a file path as source.",
    parameters={
        "type": "object",
        "properties": {
            "source": {"type": "string", "description": "Document name/path (used as identifier)"},
            "content": {"type": "string", "description": "Document text to index (if empty, reads from source path)"},
            "chunk_size": {"type": "integer", "description": "Words per chunk (default 500)"},
        },
        "required": ["source"],
    },
    handler=rag_index,
    category="rag",
)

registry.register(
    name="rag_search",
    description="Search indexed documents by semantic similarity. Returns the most relevant chunks. Use after rag_index to find information in documents.",
    parameters={
        "type": "object",
        "properties": {
            "query": {"type": "string", "description": "Search query in natural language"},
            "top_k": {"type": "integer", "description": "Number of results (default 5)"},
        },
        "required": ["query"],
    },
    handler=rag_search,
    category="rag",
)

registry.register(
    name="rag_list",
    description="List all documents indexed in the vector store for RAG.",
    parameters={"type": "object", "properties": {}, "required": []},
    handler=rag_list,
    category="rag",
)

registry.register(
    name="rag_delete",
    description="Remove a document from the RAG vector store.",
    parameters={
        "type": "object",
        "properties": {
            "source": {"type": "string", "description": "Document name/path to remove"},
        },
        "required": ["source"],
    },
    handler=rag_delete,
    category="rag",
)


# ---- Memory Tools ----

def memory_read(target: str = "memory") -> dict:
    """Read agent memory or user profile."""
    if target == "user":
        return {"target": "user", "content": memory.read_user()}
    return {"target": "memory", "content": memory.read_memory()}


registry.register(
    name="memory_read",
    description="Read persistent memory. target='memory' for agent notes, 'user' for user profile.",
    parameters={
        "type": "object",
        "properties": {
            "target": {
                "type": "string",
                "enum": ["memory", "user"],
                "description": "Which memory store to read",
            },
        },
        "required": [],
    },
    handler=memory_read,
    category="memory",
)


def memory_write(action: str, content: str, old_content: str = "", target: str = "memory", **kwargs) -> dict:
    """Write to agent memory. Actions: add, replace, remove."""
    mgr = memory
    # Smart routing: only use "user" profile for explicit user-about-me statements
    # Everything else (remember X, save this, mistakes, workflows, tasks) goes to memory
    use_user = False
    if target == "user":
        # Only allow user profile for genuine self-description patterns
        lower = content.lower()
        user_signals = ["my name is", "i am ", "i'm ", "i prefer ", "i like ", "i don't like",
                        "i work at", "i work as", "my job", "my role", "call me ",
                        "my email", "my phone", "my address", "about me"]
        use_user = any(sig in lower for sig in user_signals)

    if use_user:
        if action == "add":
            mgr.add_user(content)
        elif action == "replace":
            mgr.replace_user(old_content, content)
        elif action == "remove":
            mgr.remove_user(content)
        actual_target = "user"
    else:
        if action == "add":
            mgr.add_memory(content)
        elif action == "replace":
            mgr.replace_memory(old_content, content)
        elif action == "remove":
            mgr.remove_memory(content)
        actual_target = "memory"
    return {"status": "ok", "target": actual_target, "action": action}


registry.register(
    name="memory_write",
    description="Save information to persistent memory. target='memory' for tasks, workflows, facts, mistakes, notes, saved info, instructions (DEFAULT — use this unless saving personal info). target='user' ONLY for personal identity info (name, job title, contact info, personal preferences like 'I prefer dark mode').",
    parameters={
        "type": "object",
        "properties": {
            "target": {"type": "string", "enum": ["memory", "user"], "description": "memory = tasks, workflows, facts, mistakes, instructions (DEFAULT). user = personal identity info ONLY (name, job, contact)."},
            "action": {"type": "string", "enum": ["add", "replace", "remove"]},
            "content": {"type": "string", "description": "Content to add/remove, or new content for replace"},
            "old_content": {"type": "string", "description": "For replace: the text to find"},
        },
        "required": ["action", "content"],
    },
    handler=memory_write,
    category="memory",
)


# ---- Skill Tools ----

def skills_list() -> dict:
    """List all available skills."""
    return {"skills": skills.list_skills()}


registry.register(
    name="skills_list",
    description="List all available agent skills with names and descriptions.",
    parameters={"type": "object", "properties": {}, "required": []},
    handler=skills_list,
    category="skills",
)


def skill_view(name: str, file_path: str = "") -> dict:
    """View a skill's content."""
    return skills.view_skill(name, file_path)


registry.register(
    name="skill_view",
    description="View a skill's SKILL.md content and supporting files.",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Skill name"},
            "file_path": {"type": "string", "description": "Optional: specific file within the skill"},
        },
        "required": ["name"],
    },
    handler=skill_view,
    category="skills",
)


def skill_create(name: str, description: str, content: str, category: str = "") -> dict:
    """Create a new skill."""
    return skills.create_skill(name, description, content, category)


registry.register(
    name="skill_create",
    description="Create a new reusable skill. The content should be markdown instructions.",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Skill name (lowercase, hyphens ok)"},
            "description": {"type": "string", "description": "Short description (max 1024 chars)"},
            "content": {"type": "string", "description": "Markdown instructions for the skill"},
            "category": {"type": "string", "description": "Optional category for organization"},
        },
        "required": ["name", "description", "content"],
    },
    handler=skill_create,
    category="skills",
)


def skill_delete(name: str) -> dict:
    """Delete a skill by name."""
    return skills.delete_skill(name)


registry.register(
    name="skill_delete",
    description="Delete an existing skill by name. This is permanent.",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Skill name to delete"},
        },
        "required": ["name"],
    },
    handler=skill_delete,
    category="skills",
)


# ---- Cron Tools ----

def cron_create(name: str, command: str, interval: str = "", run_at: str = "",
                repeat: str = "once", job_type: str = "agent") -> dict:
    """Create a scheduled job. Supports interval ('5m') or time-of-day ('9:45am')."""
    print("[CRON] cron_create tool called: name=" + str(name) + " interval=" + repr(interval) + " run_at=" + repr(run_at))
    return cron.create_job(name, command, interval=interval, run_at=run_at, repeat=repeat, job_type=job_type)


registry.register(
    name="cron_create",
    description=(
        "Create a scheduled job. Two scheduling modes:\n"
        "1) interval: run repeatedly, e.g. '30s', '5m', '1h', '1d'\n"
        "2) run_at: run at a specific time, e.g. '9:45', '9:45am', '2:30pm'\n"
        "   Use repeat='daily' to run every day, or repeat='once' (default) for one-time.\n"
        "Provide EITHER interval OR run_at, not both.\n"
        "job_type='agent' sends the command as a prompt through the full Hermes agent "
        "(with all tools, memory, skills). "
        "job_type='shell' runs it as a shell command."
    ),
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Human-readable job name"},
            "command": {
                "type": "string",
                "description": "Agent prompt (type=agent) or shell command (type=shell)",
            },
            "interval": {
                "type": "string",
                "description": "How often to run, e.g. '5m', '1h', '6h', '1d'. Mutually exclusive with run_at.",
            },
            "run_at": {
                "type": "string",
                "description": "Time of day to run, e.g. '9:45', '09:45', '2:30pm'. Mutually exclusive with interval.",
            },
            "repeat": {
                "type": "string",
                "enum": ["once", "daily"],
                "description": "For run_at jobs: 'once' fires once then disables, 'daily' fires every day (default: once)",
            },
            "job_type": {
                "type": "string",
                "enum": ["agent", "shell"],
                "description": "agent = full Hermes agent loop, shell = raw command (default: agent)",
            },
        },
        "required": ["name", "command"],
    },
    handler=cron_create,
    category="cron",
)


def cron_list() -> dict:
    """List all scheduled cron jobs."""
    return cron.list_jobs()


registry.register(
    name="cron_list",
    description="List all scheduled cron jobs with their status, last run time, and results.",
    parameters={"type": "object", "properties": {}, "required": []},
    handler=cron_list,
    category="cron",
)


def cron_delete(job_id: str) -> dict:
    """Delete a scheduled cron job."""
    return cron.delete_job(job_id)


registry.register(
    name="cron_delete",
    description="Delete a scheduled cron job by its ID.",
    parameters={
        "type": "object",
        "properties": {
            "job_id": {"type": "string", "description": "The job ID to delete"},
        },
        "required": ["job_id"],
    },
    handler=cron_delete,
    category="cron",
)


def cron_toggle(job_id: str, enabled: bool) -> dict:
    """Enable or disable a cron job."""
    return cron.toggle_job(job_id, enabled)


registry.register(
    name="cron_toggle",
    description="Enable or disable a scheduled cron job.",
    parameters={
        "type": "object",
        "properties": {
            "job_id": {"type": "string", "description": "The job ID"},
            "enabled": {"type": "boolean", "description": "true to enable, false to disable"},
        },
        "required": ["job_id", "enabled"],
    },
    handler=cron_toggle,
    category="cron",
)


def cron_update(job_id: str, name: str = None, command: str = None, interval: str = None,
                run_at: str = None, repeat: str = None, job_type: str = None, enabled: bool = None) -> dict:
    """Update an existing cron job's settings."""
    kwargs = {}
    for k, v in [("name", name), ("command", command), ("interval", interval),
                 ("run_at", run_at), ("repeat", repeat), ("job_type", job_type), ("enabled", enabled)]:
        if v is not None:
            kwargs[k] = v
    return cron.update_job(job_id, **kwargs)


registry.register(
    name="cron_update",
    description=(
        "Update an existing cron job. Can change name, command, schedule (interval or run_at), "
        "repeat mode, job_type, or enabled status. Only provide the fields you want to change."
    ),
    parameters={
        "type": "object",
        "properties": {
            "job_id": {"type": "string", "description": "The job ID to update"},
            "name": {"type": "string", "description": "New job name"},
            "command": {"type": "string", "description": "New command/prompt"},
            "interval": {"type": "string", "description": "New interval (clears run_at)"},
            "run_at": {"type": "string", "description": "New time of day (clears interval)"},
            "repeat": {"type": "string", "enum": ["once", "daily"], "description": "New repeat mode for run_at jobs"},
            "job_type": {"type": "string", "enum": ["agent", "shell"], "description": "New job type"},
            "enabled": {"type": "boolean", "description": "Enable or disable the job"},
        },
        "required": ["job_id"],
    },
    handler=cron_update,
    category="cron",
)


def cron_logs(job_id: str, tail: int = 10) -> dict:
    """View recent run logs for a cron job."""
    return cron.get_job_logs(job_id, tail)


registry.register(
    name="cron_logs",
    description="View the last N run logs for a cron job, including full agent responses.",
    parameters={
        "type": "object",
        "properties": {
            "job_id": {"type": "string", "description": "The job ID"},
            "tail": {"type": "integer", "description": "Number of recent logs to return (default 10)"},
        },
        "required": ["job_id"],
    },
    handler=cron_logs,
    category="cron",
)


# ---- Package Installer (bypasses pip, works on iOS) ----

def _find_wheel_url(package_name: str, version: str = "") -> dict:
    """Query PyPI JSON API to find a compatible wheel URL."""
    if version:
        api_url = f"https://pypi.org/pypi/{package_name}/{version}/json"
    else:
        api_url = f"https://pypi.org/pypi/{package_name}/json"

    req = urllib.request.Request(api_url, headers={"User-Agent": "Pegasus/1.0"})
    with urllib.request.urlopen(req, timeout=15, context=_ssl_ctx) as resp:
        data = json.loads(resp.read().decode("utf-8", errors="replace"))

    pkg_version = data["info"]["version"]
    urls = data.get("urls", [])

    # Prefer pure-Python wheel (py3-none-any), then any wheel, then sdist
    for u in urls:
        fn = u["filename"]
        if fn.endswith(".whl") and "py3-none-any" in fn:
            return {"url": u["url"], "filename": fn, "version": pkg_version}

    for u in urls:
        fn = u["filename"]
        if fn.endswith(".whl") and ("none-any" in fn or "cp3" in fn):
            return {"url": u["url"], "filename": fn, "version": pkg_version}

    for u in urls:
        if u["filename"].endswith(".whl"):
            return {"url": u["url"], "filename": u["filename"], "version": pkg_version}

    return {"error": f"No compatible wheel found for {package_name}=={pkg_version}. Only pure-Python packages can be installed on iOS."}


def _install_single(package_name: str, version: str = "") -> dict:
    """Install a single package (no dependency resolution)."""
    # Check if already importable
    module_name = package_name.replace("-", "_").split("[")[0].lower()
    try:
        __import__(module_name)
        return {"status": "already_installed", "package": package_name}
    except ImportError:
        pass

    # Find wheel URL from PyPI
    wheel_info = _find_wheel_url(package_name, version)
    if "error" in wheel_info:
        return wheel_info

    url = wheel_info["url"]
    pkg_version = wheel_info["version"]

    # Download the wheel
    req = urllib.request.Request(url, headers={"User-Agent": "Pegasus/1.0"})
    with urllib.request.urlopen(req, timeout=60, context=_ssl_ctx) as resp:
        wheel_data = resp.read()

    # Extract wheel (it's just a zip file)
    with zipfile.ZipFile(io.BytesIO(wheel_data)) as zf:
        zf.extractall(_PACKAGES_DIR)

    return {"status": "installed", "package": package_name, "version": pkg_version}


def _get_dependencies(package_name: str) -> list:
    """Get the list of dependencies for a package from PyPI."""
    try:
        api_url = f"https://pypi.org/pypi/{package_name}/json"
        req = urllib.request.Request(api_url, headers={"User-Agent": "Pegasus/1.0"})
        with urllib.request.urlopen(req, timeout=15, context=_ssl_ctx) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="replace"))
        requires = data["info"].get("requires_dist") or []
        # Filter out extras and environment markers that don't apply
        deps = []
        for r in requires:
            # Skip optional/extra dependencies
            if "extra ==" in r or "extra==" in r:
                continue
            # Extract just the package name (before any version specifier)
            dep_name = _re.split(r'[<>=!;\[\s]', r)[0].strip()
            if dep_name:
                deps.append(dep_name)
        return deps
    except Exception:
        return []


def pip_install(package: str, version: str = "") -> dict:
    """Install a Python package from PyPI without pip or subprocess.

    Downloads the wheel from PyPI and extracts it to the packages directory.
    Automatically installs dependencies. Only pure-Python wheels work on iOS.
    """
    try:
        installed = []
        failed = []

        # Get dependencies first
        deps = _get_dependencies(package)

        # Install dependencies
        for dep in deps:
            try:
                result = _install_single(dep)
                if result.get("status") in ("installed", "already_installed"):
                    installed.append(f"{dep}: {result['status']}")
                else:
                    failed.append(f"{dep}: {result.get('error', 'unknown')}")
            except Exception as e:
                failed.append(f"{dep}: {str(e)}")

        # Install the main package
        result = _install_single(package, version)
        if "error" in result:
            return result

        # Verify import works
        module_name = package.replace("-", "_").split("[")[0].lower()
        try:
            if module_name in sys.modules:
                del sys.modules[module_name]
            __import__(module_name)
            result["import_ok"] = True
        except ImportError as e:
            result["import_ok"] = False
            result["import_error"] = str(e)

        if installed:
            result["dependencies_installed"] = installed
        if failed:
            result["dependencies_failed"] = failed
        result["location"] = _PACKAGES_DIR

        return result

    except Exception as e:
        return {"error": f"Install failed: {str(e)}"}


registry.register(
    name="pip_install",
    description="Install a Python package from PyPI. Works on iOS without pip or shell. Only pure-Python packages are supported.",
    parameters={
        "type": "object",
        "properties": {
            "package": {"type": "string", "description": "Package name (e.g. 'openpyxl', 'requests')"},
            "version": {"type": "string", "description": "Optional version (e.g. '3.1.2')"},
        },
        "required": ["package"],
    },
    handler=pip_install,
    category="system",
)


# ---- Custom Packages ----

def _list_custom_packages() -> list:
    """List all custom packages the agent has created."""
    packages = []
    if not os.path.isdir(_CUSTOM_PACKAGES_DIR):
        return packages
    for item in sorted(os.listdir(_CUSTOM_PACKAGES_DIR)):
        full = os.path.join(_CUSTOM_PACKAGES_DIR, item)
        if item.endswith(".py") and item != "__init__.py":
            name = item[:-3]
            try:
                with open(full, "r", encoding="utf-8") as f:
                    first_lines = f.read(500)
                # Extract docstring
                desc = ""
                if first_lines.startswith('"""'):
                    end = first_lines.find('"""', 3)
                    if end > 0:
                        desc = first_lines[3:end].strip().split("\n")[0]
                elif first_lines.startswith("'''"):
                    end = first_lines.find("'''", 3)
                    if end > 0:
                        desc = first_lines[3:end].strip().split("\n")[0]
                packages.append({"name": name, "type": "module", "description": desc})
            except Exception:
                packages.append({"name": name, "type": "module", "description": ""})
        elif os.path.isdir(full) and os.path.isfile(os.path.join(full, "__init__.py")):
            desc = ""
            try:
                with open(os.path.join(full, "__init__.py"), "r", encoding="utf-8") as f:
                    first_lines = f.read(500)
                if first_lines.startswith('"""'):
                    end = first_lines.find('"""', 3)
                    if end > 0:
                        desc = first_lines[3:end].strip().split("\n")[0]
            except Exception:
                pass
            packages.append({"name": item, "type": "package", "description": desc})
    return packages


def create_package(name: str, code: str, description: str = "") -> dict:
    """Create a reusable Python package that persists across sessions.

    The package is immediately importable via `import name`.
    Start the code with a docstring describing what the package does.
    """
    # Sanitize name
    safe_name = name.replace("-", "_").replace(" ", "_").lower()
    if not safe_name.isidentifier():
        return {"error": f"Invalid package name: {safe_name}. Must be a valid Python identifier."}

    # If code contains class/function definitions and is multi-module, create a package dir
    file_path = os.path.join(_CUSTOM_PACKAGES_DIR, f"{safe_name}.py")

    # Prepend docstring if description provided and code doesn't start with one
    final_code = code
    if description and not code.lstrip().startswith(('"""', "'''")):
        final_code = f'"""{description}"""\n\n{code}'

    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(final_code)

    # Clear from module cache so re-import picks up changes
    if safe_name in sys.modules:
        del sys.modules[safe_name]

    # Verify import
    try:
        __import__(safe_name)
        return {
            "status": "created",
            "name": safe_name,
            "path": file_path,
            "bytes_written": len(final_code),
            "importable": True,
        }
    except Exception as e:
        return {
            "status": "created_with_errors",
            "name": safe_name,
            "path": file_path,
            "bytes_written": len(final_code),
            "importable": False,
            "import_error": str(e),
        }


def list_custom_packages() -> dict:
    """List all custom packages created by the agent."""
    packages = _list_custom_packages()
    return {"packages": packages, "count": len(packages), "location": _CUSTOM_PACKAGES_DIR}


def delete_custom_package(name: str) -> dict:
    """Delete a custom package."""
    safe_name = name.replace("-", "_").replace(" ", "_").lower()
    file_path = os.path.join(_CUSTOM_PACKAGES_DIR, f"{safe_name}.py")
    dir_path = os.path.join(_CUSTOM_PACKAGES_DIR, safe_name)

    if os.path.isfile(file_path):
        os.remove(file_path)
        if safe_name in sys.modules:
            del sys.modules[safe_name]
        return {"status": "deleted", "name": safe_name}
    elif os.path.isdir(dir_path):
        import shutil
        shutil.rmtree(dir_path)
        if safe_name in sys.modules:
            del sys.modules[safe_name]
        return {"status": "deleted", "name": safe_name}
    return {"error": f"Package not found: {safe_name}"}


registry.register(
    name="create_package",
    description="Create a reusable Python package that persists across sessions. Immediately importable via `import name`. Use this to build utilities, helpers, and libraries you'll reuse.",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Package name (valid Python identifier, e.g. 'data_utils')"},
            "code": {"type": "string", "description": "Python source code for the package"},
            "description": {"type": "string", "description": "Short description of what the package does"},
        },
        "required": ["name", "code"],
    },
    handler=create_package,
    category="system",
)

registry.register(
    name="list_custom_packages",
    description="List all custom Python packages you have created. Shows name, type, and description.",
    parameters={
        "type": "object",
        "properties": {},
        "required": [],
    },
    handler=list_custom_packages,
    category="system",
)

registry.register(
    name="delete_custom_package",
    description="Delete a custom Python package you previously created.",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Package name to delete"},
        },
        "required": ["name"],
    },
    handler=delete_custom_package,
    category="system",
)


# ---- Background Tasks ----

import threading
import uuid as _uuid
import time as _time

class _TaskManager:
    """Manages background tasks that run independently of the main agent loop."""

    def __init__(self):
        self._tasks = {}  # id -> task dict
        self._lock = threading.Lock()

    def create(self, name, code, task_type="python"):
        """Spawn a background task. Returns task ID immediately."""
        task_id = str(_uuid.uuid4())[:8]
        task = {
            "id": task_id,
            "name": name,
            "type": task_type,
            "status": "running",
            "created": _time.strftime("%H:%M:%S"),
            "output": "",
            "error": None,
        }

        with self._lock:
            self._tasks[task_id] = task

        def _run():
            import io
            import contextlib
            stdout = io.StringIO()
            stderr = io.StringIO()
            try:
                if task_type == "python":
                    with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
                        exec(code, {"__builtins__": __builtins__})
                elif task_type == "shell":
                    from hermes_bridge.tools_builtin import shell_exec
                    result = shell_exec(code)
                    stdout.write(result.get("output", "") if isinstance(result, dict) else str(result))
                with self._lock:
                    _tlimit = 50000 if _CLOUD_MODE else 10000
                    task["output"] = stdout.getvalue()[:_tlimit]
                    task["error"] = stderr.getvalue()[:_tlimit] or None
                    task["status"] = "done"
            except Exception:
                with self._lock:
                    _tlimit = 50000 if _CLOUD_MODE else 10000
                    task["output"] = stdout.getvalue()[:_tlimit]
                    task["error"] = traceback.format_exc()[:_tlimit]
                    task["status"] = "error"

        t = threading.Thread(target=_run, daemon=True, name=f"task-{task_id}")
        t.start()
        return task_id

    def list_tasks(self):
        with self._lock:
            return [
                {"id": t["id"], "name": t["name"], "status": t["status"], "created": t["created"]}
                for t in self._tasks.values()
            ]

    def get_task(self, task_id):
        with self._lock:
            t = self._tasks.get(task_id)
            if not t:
                return {"error": f"Task not found: {task_id}"}
            return dict(t)

    def cancel_task(self, task_id):
        with self._lock:
            t = self._tasks.get(task_id)
            if not t:
                return {"error": f"Task not found: {task_id}"}
            t["status"] = "cancelled"
            return {"status": "cancelled", "id": task_id}

    def clear_done(self):
        with self._lock:
            to_remove = [k for k, v in self._tasks.items() if v["status"] in ("done", "error", "cancelled")]
            for k in to_remove:
                del self._tasks[k]
            return {"cleared": len(to_remove)}


_tasks = _TaskManager()


def task_run(name: str, code: str, task_type: str = "python") -> dict:
    """Start a background task. Returns immediately with a task ID."""
    task_id = _tasks.create(name, code, task_type)
    return {"status": "started", "task_id": task_id, "name": name}


def task_status(task_id: str = "") -> dict:
    """Check status of a task, or list all tasks if no ID given."""
    if task_id:
        return _tasks.get_task(task_id)
    return {"tasks": _tasks.list_tasks()}


def task_cancel(task_id: str) -> dict:
    """Cancel a running task."""
    return _tasks.cancel_task(task_id)


registry.register(
    name="task_run",
    description="Start a background task that runs independently. Returns a task ID. Use task_status to check results later. Good for long-running operations (web scraping, data processing, file generation) while continuing to chat.",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Short name for the task (e.g. 'scrape-prices', 'generate-report')"},
            "code": {"type": "string", "description": "Python code to run in background"},
            "task_type": {"type": "string", "enum": ["python", "shell"], "description": "Type of code (default: python)"},
        },
        "required": ["name", "code"],
    },
    handler=task_run,
    category="system",
)

registry.register(
    name="task_status",
    description="Check background task status and results. Call with no task_id to list all tasks, or with a task_id to get full output.",
    parameters={
        "type": "object",
        "properties": {
            "task_id": {"type": "string", "description": "Task ID to check (omit to list all tasks)"},
        },
        "required": [],
    },
    handler=task_status,
    category="system",
)

registry.register(
    name="task_cancel",
    description="Cancel a running background task.",
    parameters={
        "type": "object",
        "properties": {
            "task_id": {"type": "string", "description": "Task ID to cancel"},
        },
        "required": ["task_id"],
    },
    handler=task_cancel,
    category="system",
)


# ---- Motion & Sensors ----

def get_motion() -> dict:
    """Get current device motion data (pitch, roll, yaw, acceleration)."""
    return _ios_action("get_motion", {})


def get_steps(days: int = 1) -> dict:
    """Get pedometer data (steps, distance, floors climbed) for the last N days."""
    return _ios_action("get_steps", {"days": days})


def get_activity() -> dict:
    """Get current physical activity (walking, running, driving, stationary, cycling)."""
    return _ios_action("get_activity", {})


registry.register(
    name="get_motion",
    description="Get current device motion data including pitch, roll, yaw, and acceleration vectors. Useful for detecting device orientation and movement.",
    parameters={"type": "object", "properties": {}, "required": []},
    handler=get_motion,
    category="sensors",
)

registry.register(
    name="get_steps",
    description="Get pedometer data (step count, distance walked, floors climbed) for the last N days. Requires Motion & Fitness permission.",
    parameters={
        "type": "object",
        "properties": {
            "days": {"type": "integer", "description": "Number of days to look back (default 1)"},
        },
        "required": [],
    },
    handler=get_steps,
    category="sensors",
)

registry.register(
    name="get_activity",
    description="Get the user's current physical activity type (walking, running, driving, stationary, cycling). Uses Core Motion activity recognition.",
    parameters={"type": "object", "properties": {}, "required": []},
    handler=get_activity,
    category="sensors",
)


# ---- Location ----

def get_location() -> dict:
    """Get current GPS location (latitude, longitude, altitude, accuracy)."""
    return _ios_action("get_location", {}, timeout=15)


registry.register(
    name="get_location",
    description="Get the device's current GPS location including latitude, longitude, altitude, and accuracy in meters. Requires Location permission.",
    parameters={"type": "object", "properties": {}, "required": []},
    handler=get_location,
    category="location",
)


# ---- QR/Barcode Scanner ----

def scan_qr(path: str) -> dict:
    """Scan QR code or barcode from an image. Returns decoded content."""
    full = os.path.realpath(path)
    if not os.path.isfile(full):
        full = os.path.realpath(os.path.join(SANDBOX_ROOT, path))
    if not os.path.isfile(full):
        return {"error": f"Image not found: {path}"}
    return _ios_action("scan_qr", {"path": full})


registry.register(
    name="scan_qr",
    description="Scan a QR code or barcode from an image file. Returns the decoded content and barcode type. Supports QR, EAN, UPC, Code128, and more.",
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to the image file containing the QR code or barcode"},
        },
        "required": ["path"],
    },
    handler=scan_qr,
    category="vision",
)


# ---- Authentication (Face ID / Touch ID) ----

def authenticate(reason: str = "Pegasus needs to verify your identity") -> dict:
    """Authenticate user with Face ID or Touch ID. Returns success/failure."""
    return _ios_action("authenticate", {"reason": reason}, timeout=30)


registry.register(
    name="authenticate",
    description="Authenticate the user with Face ID or Touch ID. Returns success or failure. Use before sensitive operations like viewing passwords or making changes.",
    parameters={
        "type": "object",
        "properties": {
            "reason": {"type": "string", "description": "Reason shown to the user for authentication (default: 'Pegasus needs to verify your identity')"},
        },
        "required": [],
    },
    handler=authenticate,
    category="security",
)


# ---- Encryption ----

def encrypt(text: str, password: str) -> dict:
    """Encrypt text with AES-GCM using a password. Returns base64 ciphertext."""
    return _ios_action("encrypt", {"text": text, "password": password})


def decrypt(ciphertext: str, password: str) -> dict:
    """Decrypt AES-GCM ciphertext (base64) with password. Returns plaintext."""
    return _ios_action("decrypt", {"ciphertext": ciphertext, "password": password})


registry.register(
    name="encrypt",
    description="Encrypt text using AES-GCM with a password. Returns base64-encoded ciphertext. Use for storing sensitive data securely.",
    parameters={
        "type": "object",
        "properties": {
            "text": {"type": "string", "description": "Plaintext to encrypt"},
            "password": {"type": "string", "description": "Password used to derive the encryption key"},
        },
        "required": ["text", "password"],
    },
    handler=encrypt,
    category="security",
)

registry.register(
    name="decrypt",
    description="Decrypt AES-GCM ciphertext (base64-encoded) using the original password. Returns the original plaintext.",
    parameters={
        "type": "object",
        "properties": {
            "ciphertext": {"type": "string", "description": "Base64-encoded ciphertext to decrypt"},
            "password": {"type": "string", "description": "Password used during encryption"},
        },
        "required": ["ciphertext", "password"],
    },
    handler=decrypt,
    category="security",
)


# ---- Translation ----

def translate(text: str, source: str = "auto", target: str = "en") -> dict:
    """Translate text between languages using on-device Apple Translation.

    Language codes: en, es, fr, de, it, pt, zh, ja, ko, ar, ru, etc.
    Source can be 'auto' for automatic detection.
    """
    return _ios_action("translate", {"text": text, "source": source, "target": target}, timeout=15)


registry.register(
    name="translate",
    description="Translate text between languages using on-device Apple Translation. Supports en, es, fr, de, it, pt, zh, ja, ko, ar, ru, and more. Use source='auto' for automatic language detection.",
    parameters={
        "type": "object",
        "properties": {
            "text": {"type": "string", "description": "Text to translate"},
            "source": {"type": "string", "description": "Source language code, or 'auto' for detection (default 'auto')"},
            "target": {"type": "string", "description": "Target language code (default 'en')"},
        },
        "required": ["text"],
    },
    handler=translate,
    category="language",
)


# ---- Calendar Event Creation ----

def create_event(title: str, start: str, end: str = "", location: str = "", notes: str = "") -> dict:
    """Create a calendar event.

    start/end format: 'YYYY-MM-DD HH:MM' or 'YYYY-MM-DD' for all-day.
    If end is empty, creates a 1-hour event.
    """
    return _ios_action("create_event", {
        "title": title, "start": start, "end": end,
        "location": location, "notes": notes
    })


registry.register(
    name="create_event",
    description="Create a new calendar event. Supports timed events ('YYYY-MM-DD HH:MM') and all-day events ('YYYY-MM-DD'). If no end time is given, creates a 1-hour event.",
    parameters={
        "type": "object",
        "properties": {
            "title": {"type": "string", "description": "Event title"},
            "start": {"type": "string", "description": "Start date/time in 'YYYY-MM-DD HH:MM' or 'YYYY-MM-DD' for all-day"},
            "end": {"type": "string", "description": "End date/time (same format as start). Leave empty for 1-hour event."},
            "location": {"type": "string", "description": "Event location (optional)"},
            "notes": {"type": "string", "description": "Event notes/description (optional)"},
        },
        "required": ["title", "start"],
    },
    handler=create_event,
    category="calendar",
)


# ---- Contact Creation ----

def create_contact(name: str, phone: str = "", email: str = "") -> dict:
    """Create a new contact in the address book."""
    return _ios_action("create_contact", {"name": name, "phone": phone, "email": email})


registry.register(
    name="create_contact",
    description="Create a new contact in the device address book with name, phone number, and/or email address.",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string", "description": "Contact full name"},
            "phone": {"type": "string", "description": "Phone number (optional)"},
            "email": {"type": "string", "description": "Email address (optional)"},
        },
        "required": ["name"],
    },
    handler=create_contact,
    category="contacts",
)


# ---- Complete Reminder ----

def complete_reminder(title: str) -> dict:
    """Mark a reminder as completed by title."""
    return _ios_action("complete_reminder", {"title": title})


registry.register(
    name="complete_reminder",
    description="Mark a reminder as completed by its title. Searches existing reminders and marks the first match as done.",
    parameters={
        "type": "object",
        "properties": {
            "title": {"type": "string", "description": "Title of the reminder to mark as completed"},
        },
        "required": ["title"],
    },
    handler=complete_reminder,
    category="calendar",
)
